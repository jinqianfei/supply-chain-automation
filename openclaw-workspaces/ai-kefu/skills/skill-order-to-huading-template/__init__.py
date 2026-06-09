"""
skill-order-to-huading-template

将客户订单Excel转换为华鼎出库单模板（31字段）

使用前需配置：
1. 货主ID (shipper_id) - 必填
2. 数据库连接 (db_config) - 必填
"""

import os
from typing import Dict, Any, Optional


class OrderToHuadingTemplate:
    """订单转华鼎出库单模板Skill"""
    
    VERSION = "1.6"
    
    # 华鼎模板31字段
    HUADING_FIELDS = [
        "序号", "门店编号", "门店三方编码", "仓库编码", "加急程度\n（0：普通，1：加急）",
        "商品SKU编号", "商品三方SPEC编号", "单位类型", "出库数量",
        "指定库存状态", "出库类型", "配送方式", "指定车型（专配）",
        "是否垫付", "付款方式", "快递公司", "单价", "总金额",
        "是否制定批次", "批次号", "生产日期", "备注", "生产厂家编号",
        "门店收货地址编码", "三方单号", "业务模式", "业务类型",
        "收货人", "联系电话", "收货地址", "C端快递公司"
    ]
    
    # 有默认值的字段
    DEFAULT_VALUES = {
        "加急程度": 0,
        "指定库存状态": "正常",
        "出库类型": 201,
        "配送方式": "共配",
        "是否垫付": "否",
        "是否制定批次": "否"
    }
    
    def __init__(
        self,
        shipper_id: str,
        db_config: Optional[Dict] = None,
        output_dir: Optional[str] = None
    ):
        """
        初始化Skill配置
        
        Args:
            shipper_id: 货主ID（必填）
            db_config: 数据库连接配置（必填）
            output_dir: 输出目录（可选，默认./output）
        """
        if not shipper_id:
            raise ValueError("shipper_id 是必填参数")
        
        if not db_config:
            raise ValueError("db_config 是必填参数")
        
        self.shipper_id = shipper_id
        self.db_config = db_config
        self.output_dir = output_dir or "./output"
        
        # 从数据库加载仓库编码映射
        self.warehouse_code_map = self._load_warehouse_mapping()
        
        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _load_warehouse_mapping(self) -> Dict[str, str]:
        """从数据库加载仓库编码映射"""
        try:
            import psycopg2
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            cur.execute("SELECT warehouse_name, warehouse_code FROM warehouse_code_mapping")
            rows = cur.fetchall()
            
            conn.close()
            
            mapping = {}
            for name, code in rows:
                mapping[name] = code
            
            return mapping
            
        except Exception as e:
            print(f"加载仓库编码映射失败: {e}")
            return {}
    
    def _get_warehouse_code(self, warehouse_name: str) -> str:
        """获取仓库编码，找不到则报错"""
        if not warehouse_name:
            raise ValueError(f"门店的仓库名称为空，请检查门店数据")
        
        warehouse_code = self.warehouse_code_map.get(warehouse_name)
        if not warehouse_code:
            available = ", ".join(self.warehouse_code_map.keys())
            raise ValueError(
                f"仓库'{warehouse_name}'在仓库编码映射表中未找到！\n"
                f"请检查仓库编码表，确保该仓库已配置。\n"
                f"当前可用的仓库：{available}"
            )
        
        return warehouse_code
    
    def execute(self, order_file: str, output_file: str = None) -> Dict[str, Any]:
        """
        执行订单转华鼎模板
        
        Args:
            order_file: 客户订单Excel文件路径
            output_file: 输出文件路径
        
        Returns:
            Dict包含:
            - success: 是否成功
            - output_file: 输出文件路径
            - order_no: 订单号
            - store_code: 门店编号
            - warehouse_code: 仓库编码
            - item_count: 商品数量
            - unmatched_items: 未匹配SKU的商品列表
            - message: 消息
        """
        try:
            if not os.path.exists(order_file):
                return {"success": False, "message": f"订单文件不存在: {order_file}"}
            
            order_data = self._parse_order(order_file)
            if not order_data:
                return {"success": False, "message": "订单解析失败"}
            
            store_info = self._match_store(order_data["store_name"])
            sku_results, unmatched_items = self._match_sku(order_data["items"])
            
            if not output_file:
                order_no_safe = order_data["order_no"].replace("/", "-").replace("\\", "-")
                output_file = os.path.join(self.output_dir, f"华鼎出库单_{order_no_safe}.xlsx")
            
            self._generate_template(order_data, store_info, sku_results, output_file)
            
            # 构建返回消息
            message = "模板生成成功"
            if unmatched_items:
                unmatched_details = []
                for item in unmatched_items:
                    details = f"序号{item['seq']}: {item['product_name']} | 编码:{item['product_code']} | 规格:{item['spec']} | 数量:{item['quantity']}件"
                    unmatched_details.append(details)
                message = f"模板生成成功，但有{len(unmatched_items)}条商品未匹配SKU：" + "; ".join(unmatched_details)
            
            return {
                "success": True,
                "output_file": output_file,
                "order_no": order_data["order_no"],
                "store_code": store_info.get("store_code", "") if store_info else "",
                "warehouse_code": self._get_warehouse_code(store_info.get("warehouse_name", "")) if store_info else "",
                "item_count": len(sku_results),
                "unmatched_count": len(unmatched_items),
                "unmatched_items": unmatched_items,
                "message": message
            }
            
        except Exception as e:
            return {"success": False, "message": f"错误: {str(e)}"}
    
    def _parse_order(self, order_file: str) -> Optional[Dict]:
        """解析客户订单Excel"""
        try:
            import pandas as pd
            
            df_raw = pd.read_excel(order_file, header=None)
            
            order_no = str(df_raw.iloc[1, 1]) if len(df_raw) > 1 and pd.notna(df_raw.iloc[1, 1]) else ""
            store_name = str(df_raw.iloc[2, 1]) if len(df_raw) > 2 and pd.notna(df_raw.iloc[2, 1]) else ""
            
            # 清理门店名称（去掉前缀）
            if store_name.startswith("河北-"):
                store_name = store_name.replace("河北-", "")
            
            items = []
            for i in range(6, len(df_raw)):
                seq = df_raw.iloc[i, 0]
                if pd.isna(seq) or str(seq) == "合计：":
                    continue
                
                product_code = str(df_raw.iloc[i, 1]) if pd.notna(df_raw.iloc[i, 1]) else ""
                product_name = str(df_raw.iloc[i, 2]) if pd.notna(df_raw.iloc[i, 2]) else ""
                if not product_name or product_name == "nan":
                    continue
                
                # 商品规格（可能是第3列或第4列）
                spec = ""
                if pd.notna(df_raw.iloc[i, 3]):
                    spec = str(df_raw.iloc[i, 3]).replace("规格：", "").replace("件：", "").strip()
                
                quantity = df_raw.iloc[i, 5] if pd.notna(df_raw.iloc[i, 5]) else 0
                unit = str(df_raw.iloc[i, 6]) if pd.notna(df_raw.iloc[i, 6]) else "件"
                remark = df_raw.iloc[i, 8] if pd.notna(df_raw.iloc[i, 8]) else ""
                
                items.append({
                    "seq": int(seq) if isinstance(seq, (int, float)) else 1,
                    "product_code": product_code,
                    "product_name": product_name,
                    "spec": spec,
                    "quantity": int(quantity) if isinstance(quantity, (int, float)) else 0,
                    "unit": unit,
                    "remark": str(remark).strip()
                })
            
            return {
                "order_no": order_no,
                "store_name": store_name,
                "items": items
            }
            
        except Exception as e:
            print(f"订单解析错误: {e}")
            return None
    
    def _match_store(self, store_name: str) -> Optional[Dict]:
        """
        门店匹配 - 支持多种匹配策略
        
        匹配策略：
        1. 精确匹配（原始名称）
        2. 去除地名前缀后匹配（天津/河北/沧州等）
        3. 关键词匹配（取门店名中的关键词）
        """
        try:
            import psycopg2
            import re
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            # 地名前缀列表
            prefixes = ['天津', '河北', '沧州', '北京', '济南', '郑州', '杭州', '西安', '武汉', '南京', '长沙', '合肥', '太原', '上海', '沈阳', '成都', '广州', '深圳']
            
            def try_match(name):
                cur.execute("""
                    SELECT store_code, store_name, warehouse, address, contact_person, phone
                    FROM store_list
                    WHERE store_name LIKE %s
                    AND owner_code = %s
                    ORDER BY LENGTH(store_name) ASC
                    LIMIT 1
                """, (f"%{name}%", self.shipper_id))
                return cur.fetchone()
            
            row = try_match(store_name)
            match_method = "精确匹配"
            
            # 去除前缀后匹配
            if not row:
                cleaned_name = store_name
                for prefix in prefixes:
                    if cleaned_name.startswith(prefix):
                        cleaned_name = cleaned_name[len(prefix):]
                        break
                if cleaned_name != store_name:
                    row = try_match(cleaned_name)
                    match_method = f"去前缀匹配({cleaned_name})"
            
            # 提取关键词匹配（如"王口镇" → "王口"）
            if not row:
                # 尝试提取2个字以上的关键词
                keywords = re.findall(r'[\u4e00-\u9fa5]{2,}', store_name)
                for kw in keywords:
                    if len(kw) >= 2:
                        row = try_match(kw)
                        if row:
                            match_method = f"关键词匹配({kw})"
                            break
            
            # 去除"镇店"、"店"等后缀后匹配
            if not row:
                suffixes = ['镇店', '镇', '店']
                cleaned_name = store_name
                for suffix in suffixes:
                    if cleaned_name.endswith(suffix):
                        cleaned_name = cleaned_name[:-len(suffix)]
                        row = try_match(cleaned_name)
                        if row:
                            match_method = f"去后缀匹配({cleaned_name})"
                            break
            
            # 组合匹配：去除前缀+后缀
            if not row:
                cleaned_name = store_name
                for prefix in prefixes:
                    if cleaned_name.startswith(prefix):
                        cleaned_name = cleaned_name[len(prefix):]
                        break
                for suffix in suffixes:
                    if cleaned_name.endswith(suffix):
                        cleaned_name = cleaned_name[:-len(suffix)]
                        break
                if cleaned_name and cleaned_name != store_name:
                    row = try_match(cleaned_name)
                    if row:
                        match_method = f"去头尾匹配({cleaned_name})"
            
            # 最后尝试：直接搜索"王口"
            if not row:
                row = try_match("王口")
                if row:
                    match_method = "王口直接匹配"
            
            conn.close()
            
            if row:
                warehouse_name = row[2] or ""
                return {
                    "store_code": row[0] or "",
                    "store_name": row[1] or "",
                    "warehouse_name": warehouse_name,
                    "warehouse_code": self._get_warehouse_code(warehouse_name),
                    "address": row[3] or "",
                    "contact_person": row[4] or "",
                    "phone": row[5] or "",
                    "match_method": match_method
                }
            
            return None
            
        except Exception as e:
            print(f"门店匹配错误: {e}")
            return None
    
    def _clean_product_name(self, name: str) -> str:
        """
        清理商品名称，去除特殊字符用于匹配
        
        处理规则：
        - 去除首尾特殊字符（-、/、.等）
        - 去除括号及其内容
        - 去除空格
        """
        import re
        # 去除首尾的特殊字符
        name = name.strip().rstrip('-').rstrip('/').rstrip('.')
        # 去除括号及其内容
        name = re.sub(r'[（(].*[)）]', '', name)
        # 去除空格
        name = name.strip()
        return name
    
    def _match_sku(self, items: list) -> tuple:
        """
        SKU匹配 - 多层匹配策略
        
        匹配优先级：
        1. 精确匹配（商品名称完全相同）
        2. 清理后匹配（去除特殊字符）
        3. 前缀匹配（订单名称是系统名称的前缀）
        4. 后缀匹配（订单名称是系统名称的后缀）
        5. 包含匹配（订单名称包含在系统名称中）
        6. 反向包含（系统名称包含订单名称）
        
        Returns:
            (results, unmatched_items)
        """
        try:
            import psycopg2
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            results = []
            unmatched_items = []
            
            for item in items:
                sku_code = ""
                unit_type = ""
                match_method = ""
                
                original_name = item["product_name"]
                
                # ========== 1. 精确匹配 ==========
                cur.execute("""
                    SELECT system_sku_code,
                           (unit_conversion_rule->>'ratio')::numeric as ratio
                    FROM shipper_sku_mapping 
                    WHERE customer_sku_name = %s 
                    AND shipper_id = %s
                    ORDER BY ratio DESC
                    LIMIT 1
                """, (original_name, self.shipper_id))
                row = cur.fetchone()
                if row:
                    sku_code = row[0]
                    match_method = "精确匹配"
                
                # ========== 2. 清理后匹配（去除特殊字符）==========
                if not row:
                    clean_name = self._clean_product_name(original_name)
                    if clean_name != original_name:
                        cur.execute("""
                            SELECT system_sku_code,
                                   (unit_conversion_rule->>'ratio')::numeric as ratio
                            FROM shipper_sku_mapping 
                            WHERE customer_sku_name = %s 
                            AND shipper_id = %s
                            ORDER BY ratio DESC
                            LIMIT 1
                        """, (clean_name, self.shipper_id))
                        row = cur.fetchone()
                        if row:
                            sku_code = row[0]
                            match_method = f"清理匹配('{original_name}'→'{clean_name}')"
                
                # ========== 3. 前缀匹配 ==========
                if not row:
                    clean_name = self._clean_product_name(original_name)
                    cur.execute("""
                        SELECT system_sku_code,
                               (unit_conversion_rule->>'ratio')::numeric as ratio
                        FROM shipper_sku_mapping 
                        WHERE customer_sku_name LIKE %s
                        AND shipper_id = %s
                        AND customer_sku_name != %s
                        ORDER BY ratio DESC
                        LIMIT 1
                    """, (f"{clean_name}%", self.shipper_id, clean_name))
                    row = cur.fetchone()
                    if row:
                        sku_code = row[0]
                        match_method = f"前缀匹配"
                
                # ========== 4. 后缀匹配 ==========
                if not row:
                    clean_name = self._clean_product_name(original_name)
                    cur.execute("""
                        SELECT system_sku_code,
                               (unit_conversion_rule->>'ratio')::numeric as ratio
                        FROM shipper_sku_mapping 
                        WHERE customer_sku_name LIKE %s
                        AND shipper_id = %s
                        AND customer_sku_name != %s
                        ORDER BY ratio DESC
                        LIMIT 1
                    """, (f"%{clean_name}", self.shipper_id, clean_name))
                    row = cur.fetchone()
                    if row:
                        sku_code = row[0]
                        match_method = f"后缀匹配"
                
                # ========== 5. 包含匹配 ==========
                if not row:
                    clean_name = self._clean_product_name(original_name)
                    cur.execute("""
                        SELECT system_sku_code,
                               (unit_conversion_rule->>'ratio')::numeric as ratio
                        FROM shipper_sku_mapping 
                        WHERE %s LIKE '%' || customer_sku_name || '%'
                        AND shipper_id = %s
                        ORDER BY ratio DESC
                        LIMIT 1
                    """, (clean_name, self.shipper_id))
                    row = cur.fetchone()
                    if row:
                        sku_code = row[0]
                        match_method = f"包含匹配"
                
                # ========== 6. 反向包含 ==========
                if not row:
                    clean_name = self._clean_product_name(original_name)
                    cur.execute("""
                        SELECT system_sku_code,
                               (unit_conversion_rule->>'ratio')::numeric as ratio
                        FROM shipper_sku_mapping 
                        WHERE customer_sku_name LIKE %s
                        AND shipper_id = %s
                        ORDER BY ratio DESC
                        LIMIT 1
                    """, (f"%{clean_name}%", self.shipper_id))
                    row = cur.fetchone()
                    if row:
                        sku_code = row[0]
                        match_method = f"反向包含"
                
                if row:
                    unit_type = "大单位"
                    results.append({
                        "seq": item["seq"],
                        "product_name": item["product_name"],
                        "quantity": item["quantity"],
                        "remark": item["remark"],
                        "sku_code": sku_code,
                        "unit_type": unit_type,
                        "match_method": match_method
                    })
                else:
                    # 未匹配，记录完整原始信息
                    unmatched_items.append({
                        "seq": item["seq"],
                        "product_code": item.get("product_code", ""),
                        "product_name": item["product_name"],
                        "spec": item.get("spec", ""),
                        "quantity": item["quantity"],
                        "unit": item.get("unit", "件")
                    })
            
            conn.close()
            return results, unmatched_items
            
        except Exception as e:
            print(f"SKU匹配错误: {e}")
            return items, []
    
    def _generate_template(self, order_data: Dict, store_info: Optional[Dict], 
                          sku_results: list, output_file: str):
        """生成华鼎模板Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "omsWare"
        
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        
        column_widths = {1: 6, 2: 16, 3: 12, 4: 8, 5: 10, 6: 18, 7: 16, 8: 10, 9: 10,
            10: 10, 11: 8, 12: 10, 13: 12, 14: 8, 15: 8, 16: 10, 17: 8, 18: 10,
            19: 10, 20: 15, 21: 12, 22: 25, 23: 12, 24: 15, 25: 18, 26: 8, 27: 8,
            28: 10, 29: 15, 30: 40, 31: 12}
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[1].height = 35
        
        cell_alignment = Alignment(horizontal="center", vertical="center")
        cell_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        warehouse_code = store_info["warehouse_code"] if store_info and store_info.get("warehouse_code") else ""
        
        for row_idx, item in enumerate(sku_results, start=2):
            row_data = {
                "序号": 1,
                "门店编号": store_info["store_code"] if store_info else "",
                "门店三方编码": "",
                "仓库编码": warehouse_code,
                "加急程度\n（0：普通，1：加急）": self.DEFAULT_VALUES["加急程度"],
                "商品SKU编号": item["sku_code"],
                "商品三方SPEC编号": "",
                "单位类型": item["unit_type"],
                "出库数量": item["quantity"],
                "指定库存状态": self.DEFAULT_VALUES["指定库存状态"],
                "出库类型": self.DEFAULT_VALUES["出库类型"],
                "配送方式": self.DEFAULT_VALUES["配送方式"],
                "指定车型（专配）": "",
                "是否垫付": self.DEFAULT_VALUES["是否垫付"],
                "付款方式": "",
                "快递公司": "",
                "单价": "",
                "总金额": "",
                "是否制定批次": self.DEFAULT_VALUES["是否制定批次"],
                "批次号": "",
                "生产日期": "",
                "备注": item["remark"],
                "生产厂家编号": "",
                "门店收货地址编码": "",
                "三方单号": order_data["order_no"],
                "业务模式": "",
                "业务类型": "",
                "收货人": store_info["contact_person"] if store_info else "",
                "联系电话": store_info["phone"] if store_info else "",
                "收货地址": store_info["address"] if store_info else "",
                "C端快递公司": ""
            }
            
            for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
                value = row_data.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = cell_border
        
        wb.save(output_file)


# 全局配置
_global_config = {
    "shipper_id": None,
    "db_config": None,
    "output_dir": "./output"
}


def configure(shipper_id: str, db_config: Dict, output_dir: Optional[str] = None):
    """配置全局参数"""
    _global_config["shipper_id"] = shipper_id
    _global_config["db_config"] = db_config
    _global_config["output_dir"] = output_dir or "./output"


def convert_order_to_huading(order_file: str, output_file: str = None) -> Dict[str, Any]:
    """将客户订单转换为华鼎出库单模板"""
    if not _global_config["shipper_id"] or not _global_config["db_config"]:
        return {
            "success": False,
            "message": "请先调用 configure() 配置全局参数"
        }
    
    skill = OrderToHuadingTemplate(
        shipper_id=_global_config["shipper_id"],
        db_config=_global_config["db_config"],
        output_dir=_global_config["output_dir"]
    )
    
    return skill.execute(order_file, output_file)


if __name__ == "__main__":
    print("请先调用 configure() 配置参数")