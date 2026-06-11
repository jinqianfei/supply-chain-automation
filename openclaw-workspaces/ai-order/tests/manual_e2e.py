#!/usr/bin/env python3
"""
manual_e2e.py — 手动端到端测试脚本（金姐专用）

完整跑一遍 skill_order_to_huading_template v5.11.2 全流程：
  Excel 解析 → 门店匹配 → 用户确认 → SKU 匹配 → 31 字段 Excel 生成

使用：
  cd /Users/jinqianfei/openclaw-workspaces/ai-order
  python3 tests/manual_e2e.py [订单编号]
  
订单编号：
  1 = 洪洪通_1店1项.xlsx (1店1项，简单)
  2 = 天津仓_2店11项.xlsx (2店11项，复杂)
  (无参数) = 1
"""
import sys
import os
import json
from pathlib import Path

# ===== 路径设置 =====
WORKSPACE = Path("/Users/jinqianfei/openclaw-workspaces/ai-order")
SKILL_DIR = WORKSPACE / "skills" / "skill_order_to_huading_template"
sys.path.insert(0, str(SKILL_DIR))
sys.path.insert(0, str(WORKSPACE))

# 强制 working directory 是 workspace（确保 .env 找到）
os.chdir(WORKSPACE)


def banner(msg: str):
    """打印大标题"""
    print("\n" + "=" * 70)
    print(f"  {msg}")
    print("=" * 70)


def step(msg: str):
    """打印步骤"""
    print(f"\n→ {msg}")


def show_dict(d: dict, indent: int = 2, max_v_len: int = 80):
    """打印 dict（截断长字符串）"""
    def _trunc(v):
        s = str(v)
        return s[:max_v_len] + "..." if len(s) > max_v_len else s
    print(json.dumps({k: _trunc(v) for k, v in d.items()}, 
                     indent=indent, ensure_ascii=False, default=str))


# ===== 主流程 =====
def run(order_no: int = 1):
    banner(f"手动端到端测试 — 订单 #{order_no} ({'洪洪通 1店1项' if order_no == 1 else '天津仓 2店11项'})")
    
    # ===== 1. 加载 skill =====
    step("1. 加载 skill v5.11.2")
    from skills.skill_order_to_huading_template import OrderToHuadingTemplate
    from db.connection import get_default_db_config
    
    db_config = get_default_db_config()
    print(f"   DB host: {db_config['host'][:30]}...")
    print(f"   DB name: {db_config['database']}")
    
    skill = OrderToHuadingTemplate(
        db_config=db_config,
        output_dir=str(WORKSPACE / "test_output")
    )
    print(f"   ✅ skill 加载成功 (version={skill.VERSION})")
    print(f"   仓库编码映射: {len(skill.warehouse_code_map)} 条")
    
    # ===== 2. 选择订单 =====
    step(f"2. 准备订单文件（订单 #{order_no}）")
    if order_no == 1:
        order_file = WORKSPACE / "data" / "test_orders" / "洪洪通_1店1项.xlsx"
    elif order_no == 2:
        order_file = WORKSPACE / "data" / "test_orders" / "天津仓_2店11项.xlsx"
    else:
        print(f"   ❌ 不支持的订单编号: {order_no}")
        return
    
    if not order_file.exists():
        print(f"   ❌ 文件不存在: {order_file}")
        return
    print(f"   文件: {order_file}")
    print(f"   大小: {order_file.stat().st_size} bytes")
    
    # ===== 3. 第 1 次 execute：不传 confirmed_store，触发门店确认 =====
    step("3. 第 1 次 execute（不传 confirmed_store）— 预期返回 need_store_confirm")
    print("   调 skill.execute(order_input=...)")
    
    result_1 = skill.execute(order_input=str(order_file))
    
    print(f"\n   返回字段:")
    for k in result_1.keys():
        v = result_1.get(k)
        if isinstance(v, str) and len(v) > 60:
            v = v[:60] + "..."
        print(f"     {k}: {v}")
    
    if not result_1.get("need_store_confirm"):
        print(f"\n   ⚠️  期望 need_store_confirm=True，实际是 {result_1.get('need_store_confirm')}")
        if result_1.get("success"):
            print("   说明：单条订单已自动确认（exact match）")
        return result_1
    
    # ===== 4. 展示门店候选 =====
    step("4. 展示门店匹配候选")
    matched = result_1.get("matched_store", {})
    candidates = result_1.get("candidates", [])
    
    print(f"\n   提交门店名: {result_1.get('store_name_submitted')}")
    print(f"\n   最佳匹配: {matched.get('store_name')} (相似度={matched.get('similarity'):.2%}, type={matched.get('match_type')})")
    print(f"   候选数: {len(candidates)}")
    
    for i, c in enumerate(candidates[:5]):
        print(f"\n   [{i}] {c.get('store_name')}")
        print(f"        store_code:  {c.get('store_code')}")
        print(f"        owner_code:  {c.get('owner_code')}")
        print(f"        match_type:  {c.get('match_type')}")
        print(f"        match_method: {c.get('match_method')}")
        print(f"        similarity:  {c.get('similarity')}")
    
    # ===== 5. 用户确认 =====
    step("5. 用户确认门店（手动选择）")
    print("   选项：输入候选编号 (0/1/2/...) 或按 ENTER 用最佳匹配 [0]")
    
    try:
        user_input = input("   选哪个门店？ [0]: ").strip()
        idx = int(user_input) if user_input else 0
        if idx < 0 or idx >= len(candidates):
            print(f"   ⚠️  编号 {idx} 越界，用 [0]")
            idx = 0
    except (ValueError, EOFError):
        idx = 0
    
    chosen = candidates[idx]
    print(f"\n   ✅ 已选: [{idx}] {chosen.get('store_name')} (owner={chosen.get('owner_code')})")
    
    # ===== 6. 第 2 次 execute：传 confirmed_store =====
    step("6. 第 2 次 execute（传 confirmed_store）— 预期生成 Excel")
    print(f"   调 skill.execute(order_input=..., confirmed_store=...)")
    
    result_2 = skill.execute(
        order_input=str(order_file),
        confirmed_store=chosen
    )
    
    # ===== 7. 展示结果 =====
    step("7. 最终结果")
    print(f"\n   success: {result_2.get('success')}")
    print(f"   has_issues: {result_2.get('has_issues')}")
    print(f"   output_file: {result_2.get('output_file')}")
    print(f"   file_name: {result_2.get('file_name')}")
    print(f"   order_no: {result_2.get('order_no')}")
    print(f"   store_names: {result_2.get('store_names')}")
    print(f"   store_count: {result_2.get('store_count')}")
    print(f"   item_count: {result_2.get('item_count')}")
    print(f"   matched_count: {result_2.get('matched_count')}")
    print(f"   unmatched_count: {result_2.get('unmatched_count')}")
    print(f"   unmatched_items: {result_2.get('unmatched_items')}")
    print(f"   message: {result_2.get('message')}")
    
    # ===== 8. SKU 映射明细 =====
    step("8. SKU 映射对照表（详细）")
    review = result_2.get("review_data", {})
    if review:
        print(f"\n   summary: {review.get('summary')}")
        print(f"\n   mapping_table 长度: {len(review.get('mapping_table', []))}")
        for row in review.get("mapping_table", [])[:15]:
            print(f"     {row}")
        if len(review.get("mapping_table", [])) > 15:
            print(f"     ... (还有 {len(review.get('mapping_table', []))-15} 行)")
        
        # 告警
        alerts = review.get("alerts", [])
        if alerts:
            print(f"\n   ⚠️  告警 ({len(alerts)} 条):")
            for a in alerts[:5]:
                print(f"     {a}")
    
    # ===== 9. Excel 验证 =====
    step("9. 验证生成的 Excel")
    output_file = result_2.get("output_file")
    if output_file and os.path.exists(output_file):
        import openpyxl
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        print(f"\n   ✅ Excel 存在: {output_file}")
        print(f"   Sheet: {ws.title}")
        print(f"   行数: {ws.max_row} (1 表头 + {ws.max_row-1} 数据)")
        print(f"   列数: {ws.max_column}")
        print(f"\n   表头 (前 10 列): {[ws.cell(1, c).value for c in range(1, 11)]}")
        if ws.max_row >= 2:
            print(f"   第 1 数据行: {[ws.cell(2, c).value for c in range(1, 11)]}")
    else:
        print(f"   ❌ Excel 未生成: {output_file}")
    
    banner("✅ 手动测试完成")
    return result_2


if __name__ == "__main__":
    order_no = int(sys.argv[1]) if len(sys.argv) > 1 else 1
    run(order_no)
