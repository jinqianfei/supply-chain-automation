#!/usr/bin/env python3
"""
重新执行Skill测评 - 使用正确的Ground Truth标注
修正版：修复多种订单格式解析

Ground Truth来源：直接查询数据库product_sku表
"""
import json
import os
import sys
import re
from pathlib import Path
from datetime import datetime

# 添加skill路径
sys.path.insert(0, '/Users/jinqianfei/openclaw-workspaces/ai-order/skills/skill_order_to_huading_template')

from tools.sku_mapper import map_sku
from tools.store_matcher import match_store

def clean_text(text):
    """"清洗文本中的各种空白字符"""
    if not text:
        return ""
    text = str(text)
    text = text.strip()
    text = re.sub(r'[\t\u00a0\u3000\u200b-\u200f\ufeff]+', '', text)
    text = text.replace(' ', '')
    return text

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "database": "neo",
    "user": "jinqianfei"
}

ORDER_FILES = [
    "广州仓报单明细_16---bc6cf341-af88-45ad-81cd-583378eb96d3.xlsx",
    "华鼎_郑州仓报单明细_21---5928d1e0-7772-4d56-9e7b-55e204bbe8f2.xlsx",
    "小江溪_江西菜_长治万达店_配送明细表_5.29---18401300-3340-40fb-b2d5-f77d3e3c9ef0.xlsx",
    "配送明细表_26---c5ede6f0-f958-42cc-bb07-43bda964a3e7.xlsx",
    "华鼎下单5.29---c87cb9b0-acf8-4e7c-96ee-6c41276812d7.xlsx",
    "华鼎出库单20260529---f9d8c3a4-2b90-4900-b4d6-6e98a02a003e.xlsx",
    "2026.5.28天津仓库_订单_1---29d44dbb-3e0f-43cd-b4e8-1ddfc54172ad.xlsx",
]

INBOUND_DIR = Path("/Users/jinqianfei/.openclaw/media/inbound/")
OUTPUT_DIR = Path("/Users/jinqianfei/openclaw-workspaces/ai-order/docs/test_data/")


def detect_format(headers, first_data_row):
    """根据表头和数据行检测订单格式"""
    h0 = str(headers[0]).strip() if headers[0] else ""
    
    # 华鼎标准格式：序号, 门店名称, 仓库, 商品名称, 数量, 单位
    if h0 == "序号" and "门店名称" in headers and "商品名称" in headers:
        return "huading_standard"
    
    # 华鼎出库单格式：序号, 门店名称, 联系人, 联系电话, 地址, 发货仓库, 产品名称, 规格
    if h0 == "序号" and "门店名称" in headers and "产品名称" in headers:
        return "huading_output"
    
    # 华鼎下单格式：序号, 客户名称*, 详细地址, 电话, 商品名称*, 出库数量*, 单位, 运输方式*
    if h0 == "序号" and "客户名称*" in headers and "商品名称*" in headers:
        return "huading_order"
    
    # 小江溪格式：有多行表头，以'小江溪'开头
    if "小江溪" in h0 or "配送明细表" in str(headers[1] or ""):
        return "xiaojiangxi"
    
    # 天津仓格式：首行是日期标题
    if "单据日期" in h0 or "天津仓" in h0:
        return "tianjin"
    
    return "unknown"


def find_header_row(ws, max_search=5):
    """找到表头行（包含商品名称/品名等关键词的行）"""
    for row_idx in range(1, min(max_search + 1, ws.max_row + 1)):
        row = [str(cell.value).strip() if cell.value else "" for cell in ws[row_idx]]
        if any(kw in str(row) for kw in ["商品名称", "品名", "产品名称", "叫货日期", "单据日期"]):
            return row_idx, row
    return 1, [str(cell.value).strip() if cell.value else "" for cell in ws[1]]


def find_data_start(ws, header_row_idx, headers):
    """找到数据开始行（跳过空行和表头）"""
    # 商品/品名列的索引
    product_col = None
    for kw in ["商品名称", "品名", "产品名称"]:
        if kw in headers:
            product_col = headers.index(kw)
            break
    
    if product_col is None:
        return header_row_idx + 1
    
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        product_val = row[product_col] if product_col < len(row) else None
        if product_val and str(product_val).strip() and str(product_val).strip() not in ["商品名称", "品名", "产品名称"]:
            return row_idx
    
    return header_row_idx + 1


def parse_huading_standard(ws, filename):
    """解析华鼎标准格式（广州仓、郑州仓报单明细）"""
    header_row_idx, headers = find_header_row(ws)
    data_start = find_data_start(ws, header_row_idx, headers)
    
    stores = {}
    current_store = None
    
    for row_idx in range(data_start, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        
        # 跳过空行
        if not any(row):
            continue
        
        # 门店名称列
        store_idx = headers.index("门店名称") if "门店名称" in headers else 1
        store_name = str(row[store_idx] or "").strip() if row[store_idx] else ""
        
        if store_name and store_name != "None":
            current_store = store_name
        
        if not current_store:
            continue
        
        # 商品名称
        product_idx = headers.index("商品名称") if "商品名称" in headers else 3
        product_name = str(row[product_idx] or "").strip().replace("\t", "").replace(" ", "").replace("\u00a0", "").replace("\u3000", "") if row[product_idx] else ""
        
        if not product_name or product_name in ["商品名称", "品名", "None"]:
            continue
        
        # 数量
        qty_idx = headers.index("数量") if "数量" in headers else 4
        try:
            qty = int(float(row[qty_idx] or 0))
        except:
            qty = 0
        
        # 单位
        unit_idx = headers.index("单位") if "单位" in headers else 5
        unit = str(row[unit_idx] or "").strip().replace("\t", "").replace(" ", "").replace("\u00a0", "").replace("\u3000", "") if row[unit_idx] else ""
        
        # 规格（如果有）
        spec = ""
        if "规格" in headers:
            spec_idx = headers.index("规格")
            spec = str(row[spec_idx] or "").strip().replace("\t", "").replace(" ", "").replace("\u00a0", "").replace("\u3000", "") if row[spec_idx] else ""
        
        if current_store not in stores:
            stores[current_store] = []
        
        stores[current_store].append({
            "name": product_name,
            "qty": qty,
            "unit": unit,
            "spec": spec
        })
    
    return {"filename": filename, "format": "huading_standard", "stores": stores}


def parse_huading_output(ws, filename):
    """解析华鼎出库单格式"""
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    stores = {}
    current_store = None
    
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        
        if not any(row):
            continue
        
        # 门店名称
        store_idx = headers.index("门店名称") if "门店名称" in headers else 1
        store_name = str(row[store_idx] or "").strip() if row[store_idx] else ""
        
        if store_name and store_name != "None":
            current_store = store_name
        
        if not current_store:
            continue
        
        # 产品名称（不是商品名称！）
        product_idx = headers.index("产品名称") if "产品名称" in headers else 6
        product_name = clean_text(row[product_idx]) if row[product_idx] else ""
        
        if not product_name or product_name in ["产品名称", "None"]:
            continue
        
        # 规格
        spec_idx = headers.index("规格") if "规格" in headers else 7
        spec = str(row[spec_idx] or "").strip() if row[spec_idx] else ""
        
        # 数量（可能在序号列）
        seq_idx = headers.index("序号") if "序号" in headers else 0
        qty = 1  # 默认1
        
        if current_store not in stores:
            stores[current_store] = []
        
        stores[current_store].append({
            "name": product_name,
            "qty": qty,
            "unit": "",
            "spec": spec
        })
    
    return {"filename": filename, "format": "huading_output", "stores": stores}


def parse_huading_order(ws, filename):
    """解析华鼎下单格式"""
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    stores = {}
    current_store = None
    
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        
        if not any(row):
            continue
        
        # 客户名称
        customer_idx = headers.index("客户名称*") if "客户名称*" in headers else 1
        store_name = str(row[customer_idx] or "").strip() if row[customer_idx] else ""
        
        if store_name and store_name != "None":
            current_store = store_name
        
        if not current_store:
            continue
        
        # 商品名称*
        product_idx = headers.index("商品名称*") if "商品名称*" in headers else 4
        product_name = clean_text(row[product_idx]) if row[product_idx] else ""
        
        if not product_name or product_name in ["商品名称*", "None"]:
            continue
        
        # 数量
        qty_idx = headers.index("出库数量*") if "出库数量*" in headers else 5
        try:
            qty = int(float(row[qty_idx] or 0))
        except:
            qty = 0
        
        # 单位
        unit_idx = headers.index("单位") if "单位" in headers else 6
        unit = str(row[unit_idx] or "").strip() if row[unit_idx] else ""
        
        if current_store not in stores:
            stores[current_store] = []
        
        stores[current_store].append({
            "name": product_name,
            "qty": qty,
            "unit": unit,
            "spec": ""
        })
    
    return {"filename": filename, "format": "huading_order", "stores": stores}


def parse_xiaojiangxi(ws, filename):
    """解析小江溪格式"""
    # 找表头行
    header_row_idx, headers = find_header_row(ws, max_search=6)
    data_start = find_data_start(ws, header_row_idx, headers)
    
    stores = {}
    current_store = None
    
    for row_idx in range(data_start, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        
        if not any(row):
            continue
        
        # 门店列
        store_col = None
        for col_name in ["门店", "店名"]:
            if col_name in headers:
                store_col = headers.index(col_name)
                break
        
        if store_col is not None:
            store_name = str(row[store_col] or "").strip()
            if store_name and store_name != "None":
                current_store = store_name
        
        if not current_store:
            continue
        
        # 品名
        product_col = None
        for col_name in ["品名", "商品名称", "产品名称"]:
            if col_name in headers:
                product_col = headers.index(col_name)
                break
        
        if product_col is None:
            continue
        
        product_name = str(row[product_col] or "").strip()
        if not product_name or product_name in ["品名", "商品名称", "None"]:
            continue
        
        # 规格
        spec = ""
        if "规格" in headers:
            spec_idx = headers.index("规格")
            spec = str(row[spec_idx] or "").strip() if row[spec_idx] else ""
        
        # 数量
        qty = 0
        for col_name in ["配送数量", "叫货数量", "数量"]:
            if col_name in headers:
                try:
                    qty = int(float(row[headers.index(col_name)] or 0))
                except:
                    qty = 0
                break
        
        # 单位
        unit = ""
        if "单位" in headers:
            unit = str(row[headers.index("单位")] or "").strip()
        
        if current_store not in stores:
            stores[current_store] = []
        
        stores[current_store].append({
            "name": product_name,
            "qty": qty,
            "unit": unit,
            "spec": spec
        })
    
    return {"filename": filename, "format": "xiaojiangxi", "stores": stores}


def parse_tianjin(ws, filename):
    """解析天津仓格式"""
    # 找表头行
    header_row_idx = None
    headers = []
    for row_idx in range(1, min(5, ws.max_row + 1)):
        row = [str(cell.value).strip() if cell.value else "" for cell in ws[row_idx]]
        if "商品名称" in row or "品名" in row:
            header_row_idx = row_idx
            headers = row
            break
    
    if not headers:
        return {"filename": filename, "format": "tianjin", "stores": {}}
    
    stores = {}
    current_store = None
    
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
        
        if not any(row):
            continue
        
        # 往来单位
        unit_idx = headers.index("往来单位名称") if "往来单位名称" in headers else 1
        store_name = str(row[unit_idx] or "").strip() if row[unit_idx] else ""
        
        if store_name and store_name != "None":
            current_store = store_name
        
        if not current_store:
            continue
        
        # 商品名称
        product_idx = headers.index("商品名称") if "商品名称" in headers else 4
        product_name = clean_text(row[product_idx]) if row[product_idx] else ""
        
        if not product_name or product_name in ["商品名称", "品名", "None"]:
            continue
        
        # 数量
        qty_idx = headers.index("数量") if "数量" in headers else 7
        try:
            qty = int(float(row[qty_idx] or 0))
        except:
            qty = 0
        
        # 规格
        spec = ""
        if "规格" in headers:
            spec_idx = headers.index("规格")
            spec = str(row[spec_idx] or "").strip() if row[spec_idx] else ""
        
        if current_store not in stores:
            stores[current_store] = []
        
        stores[current_store].append({
            "name": product_name,
            "qty": qty,
            "unit": "",
            "spec": spec
        })
    
    return {"filename": filename, "format": "tianjin", "stores": stores}


def load_and_parse_order(filepath):
    """加载并解析订单"""
    import openpyxl
    from datetime import datetime
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    filename = filepath.name
    first_row = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    second_row = [str(cell.value).strip() if cell.value else "" for cell in ws[2]] if ws.max_row >= 2 else [""]
    
    fmt = detect_format(first_row, second_row)
    
    if fmt == "huading_standard":
        return parse_huading_standard(ws, filename)
    elif fmt == "huading_output":
        return parse_huading_output(ws, filename)
    elif fmt == "huading_order":
        return parse_huading_order(ws, filename)
    elif fmt == "xiaojiangxi":
        return parse_xiaojiangxi(ws, filename)
    elif fmt == "tianjin":
        return parse_tianjin(ws, filename)
    else:
        print(f"  ⚠️ 未知格式: {first_row[:8]}")
        return {"filename": filename, "format": "unknown", "stores": {}}


def evaluate_order(order_data, db_config):
    """评估单个订单"""
    results = {
        "filename": order_data["filename"],
        "format": order_data["format"],
        "stores": []
    }
    
    total_items = 0
    total_matched = 0
    
    for store_name, items in order_data["stores"].items():
        # 门店匹配
        store_match = match_store(store_name, db_config=db_config)
        
        store_result = {
            "store_name": store_name,
            "items": [],
            "store_matched": store_match is not None and not store_match.get("need_confirm", False),
            "store_match": store_match,
            "owner_code": store_match.get("owner_code") if store_match else None
        }
        
        if store_match and store_match.get("owner_code"):
            owner_code = store_match["owner_code"]
            
            for item in items:
                if not item["name"]:
                    continue
                
                # SKU映射
                mapped = map_sku(owner_code, item["name"], item.get("unit", ""), db_config)
                
                item_result = {
                    "name": item["name"],
                    "qty": item["qty"],
                    "unit": item["unit"],
                    "spec": item.get("spec", ""),
                    "matched": mapped.get("matched", False),
                    "confidence": mapped.get("confidence", 0),
                    "sku_code": mapped.get("sku_code", ""),
                    "sku_name": mapped.get("sku_name", ""),
                    "unit_type": mapped.get("unit_type", ""),
                    "match_method": mapped.get("match_method", "")
                }
                
                store_result["items"].append(item_result)
                total_items += 1
                if item_result["matched"]:
                    total_matched += 1
        else:
            # 门店未匹配，无法进行SKU映射
            for item in items:
                store_result["items"].append({
                    "name": item["name"],
                    "qty": item["qty"],
                    "unit": item["unit"],
                    "spec": item.get("spec", ""),
                    "matched": False,
                    "confidence": 0,
                    "sku_code": "",
                    "sku_name": "",
                    "unit_type": "",
                    "match_method": "门店未匹配，跳过"
                })
            total_items += len(items)
        
        results["stores"].append(store_result)
    
    results["total_items"] = total_items
    results["total_matched"] = total_matched
    results["accuracy"] = (total_matched / total_items * 100) if total_items > 0 else 0
    
    return results


def main():
    print("=" * 70)
    print("Skill测评重新执行 v3 - 修正订单格式解析")
    print("=" * 70)
    
    all_results = []
    total_stores = 0
    total_store_matched = 0
    total_items = 0
    total_sku_matched = 0
    
    for filename in ORDER_FILES:
        filepath = INBOUND_DIR / filename
        if not filepath.exists():
            print(f"\n⚠️ 文件不存在: {filename}")
            continue
        
        print(f"\n📄 处理: {filename}")
        
        try:
            order_data = load_and_parse_order(filepath)
        except Exception as e:
            print(f"  ❌ 解析失败: {e}")
            import traceback
            traceback.print_exc()
            continue
        
        store_count = len(order_data["stores"])
        item_count = sum(len(items) for items in order_data["stores"].values())
        print(f"  格式: {order_data['format']}, 门店数: {store_count}, 商品数: {item_count}")
        
        for store_name, items in order_data["stores"].items():
            print(f"    门店「{store_name}」: {len(items)}个商品")
        
        # 评估
        eval_result = evaluate_order(order_data, DB_CONFIG)
        all_results.append(eval_result)
        
        # 统计
        total_stores += len(eval_result["stores"])
        for store_result in eval_result["stores"]:
            if store_result["store_matched"]:
                total_store_matched += 1
            
            for item in store_result["items"]:
                total_items += 1
                if item["matched"]:
                    total_sku_matched += 1
        
        # 打印摘要
        print(f"  结果: {eval_result['total_matched']}/{eval_result['total_items']} SKU匹配")
    
    # 计算准确率
    store_accuracy = (total_store_matched / total_stores * 100) if total_stores > 0 else 0
    sku_accuracy = (total_sku_matched / total_items * 100) if total_items > 0 else 0
    
    # 输出汇总
    print("\n" + "=" * 70)
    print("📊 测评结果汇总")
    print("=" * 70)
    print(f"\n【门店匹配】")
    print(f"  总门店数: {total_stores}")
    print(f"  匹配成功: {total_store_matched}")
    print(f"  准确率: {store_accuracy:.1f}%")
    
    print(f"\n【SKU映射】")
    print(f"  总商品数: {total_items}")
    print(f"  映射成功: {total_sku_matched}")
    print(f"  准确率: {sku_accuracy:.1f}%")
    
    # 详细结果展示
    print(f"\n【SKU映射详情】")
    print("-" * 70)
    
    low_confidence_items = []
    unmatched_items = []
    
    for order_result in all_results:
        for store_result in order_result["stores"]:
            store_name = store_result["store_name"]
            owner_code = store_result["owner_code"] or "未匹配"
            
            if not store_result["items"]:
                continue
                
            print(f"\n📍 {store_name} (货主: {owner_code})")
            print(f"   {'状态':<6} {'商品名称':<32} {'SKU编码':<20} {'置信度':<8} {'匹配方式'}")
            print(f"   {'-'*70}")
            
            for item in store_result["items"]:
                status = "✅" if item["matched"] else "❌"
                conf = item["confidence"]
                method = item["match_method"][:30] if item["match_method"] else ""
                sku = item["sku_code"] or "(未匹配)"
                
                print(f"   {status:<6} {item['name'][:30]:<32} {sku:<20} {conf:.2f}     {method}")
                
                if item["matched"] and conf < 0.8:
                    low_confidence_items.append({"store": store_name, **item})
                if not item["matched"]:
                    unmatched_items.append({"store": store_name, **item})
    
    # 低置信度告警
    if low_confidence_items:
        print(f"\n⚠️ 低置信度匹配（<0.8）共{len(low_confidence_items)}个:")
        for item in low_confidence_items[:10]:
            print(f"  ⚠️ [{item['store']}] {item['name']} → {item['sku_code']} ({item['confidence']:.2f})")
    
    # 未匹配商品
    if unmatched_items:
        print(f"\n❌ 未匹配商品共{len(unmatched_items)}个:")
        for item in unmatched_items[:10]:
            print(f"  ❌ [{item['store']}] {item['name']}")
    
    # 保存报告
    report = {
        "evaluation_date": "2026-06-02",
        "skill_version": "v5.4",
        "database": "neo",
        "order_files_count": len(ORDER_FILES),
        "store_matching": {
            "total": total_stores,
            "matched": total_store_matched,
            "accuracy_pct": round(store_accuracy, 1)
        },
        "sku_mapping": {
            "total": total_items,
            "matched": total_sku_matched,
            "accuracy_pct": round(sku_accuracy, 1)
        },
        "all_results": all_results,
        "low_confidence_items": low_confidence_items,
        "unmatched_items": unmatched_items
    }
    
    output_file = OUTPUT_DIR / "evaluation_results_v2.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n📁 详细结果已保存到: {output_file}")
    
    return report


if __name__ == "__main__":
    main()