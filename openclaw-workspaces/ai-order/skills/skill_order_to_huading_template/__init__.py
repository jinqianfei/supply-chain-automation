"""
skill-order-to-huading-template

将客户订单Excel转换为华鼎出库单模板（31字段）
支持Excel/图片/PDF/文字多种输入格式

使用前需配置：
1. 数据库连接 (db_config) - 必填
"""

import os
import re
import datetime
import time
import uuid
from typing import Dict, Any, Optional, Union, List, Tuple

# ── v5.9.0 Phase 1：事件总线 + 反馈采集器（懒加载，单例）────────
try:
    from events.bus import EventBus
    from learn.collector import init_feedback_collector, get_feedback_collector
    _HAS_EVENT_BUS = True
except ImportError:
    _HAS_EVENT_BUS = False
    EventBus = None
    init_feedback_collector = None
    get_feedback_collector = None
class OrderSkillError(Exception):
    """订单映射Skill专用异常"""
    def __init__(self, code: str, message: str, detail: str = ""):
        self.code = code
        self.message = message
        self.detail = detail
        super().__init__(f"[{code}] {message}")


ERROR_CODES = {
    # 文件相关
    "E001": {"title": "文件不存在", "message": "您上传的文件未找到，请检查后重新上传"},
    "E002": {"title": "文件格式错误", "message": "不支持的文件格式，请上传Excel、图片或PDF文件"},
    "E003": {"title": "文件读取失败", "message": "文件读取失败，可能是文件损坏，请重新上传"},
    
    # 订单解析相关
    "E101": {"title": "订单解析失败", "message": "无法识别订单格式，请检查订单内容或联系客服"},
    "E102": {"title": "缺少必填字段", "message": "订单中缺少必要的商品信息，请检查后重新上传"},
    "E103": {"title": "需要OCR识别", "message": "图片订单需要识别，请稍候..."},
    
    # 门店匹配相关
    "E201": {"title": "门店未找到", "message": "未找到匹配门店，请确认门店名称是否正确"},
    "E202": {"title": "门店匹配需要确认", "message": "找到多个相似门店，请选择正确的门店"},
    
    # SKU映射相关
    "E301": {"title": "SKU映射失败", "message": "商品匹配失败，请检查商品名称或联系客服"},
    "E302": {"title": "商品不存在", "message": "该商品在系统中未找到，请确认商品名称"},
    
    # 数据库相关
    "E401": {"title": "数据库连接失败", "message": "系统暂时无法处理订单，请稍后再试"},
    "E402": {"title": "数据库查询失败", "message": "查询数据时出错，请稍后再试"},
    
    # 通用
    "E999": {"title": "未知错误", "message": "系统繁忙，请稍后再试"},
}


def get_friendly_error(code: str, detail: str = "") -> Dict[str, Any]:
    """
    将错误码转换为用户友好的错误信息
    
    Args:
        code: 错误码（如 E001, E101）
        detail: 详细信息（用于日志或调试）
    
    Returns:
        Dict: {success, error_code, error_title, error_message}
    """
    err_info = ERROR_CODES.get(code, ERROR_CODES["E999"])
    return {
        "success": False,
        "error_code": code,
        "error_title": err_info["title"],
        "error_message": err_info["message"],
        "_debug_detail": detail if os.getenv("DEBUG") else ""  # 仅调试模式显示
    }


def wrap_error(e: Exception, code: str = "E999") -> Dict[str, Any]:
    """
    将异常包装为友好错误返回
    
    Args:
        e: 原始异常
        code: 错误码
    
    Returns:
        Dict: 友好的错误信息字典
    """
    detail = str(e)
    
    # 特殊异常处理
    if "psycopg2" in str(type(e)):
        return get_friendly_error("E401", detail)
    elif "ConnectionRefused" in str(e):
        return get_friendly_error("E401", detail)
    elif "FileNotFoundError" in str(type(e)):
        return get_friendly_error("E001", detail)
    elif "KeyError" in str(type(e)):
        return get_friendly_error("E102", detail)
    elif "IndexError" in str(type(e)):
        return get_friendly_error("E102", detail)
    elif "UnicodeDecodeError" in str(type(e)):
        return get_friendly_error("E003", detail)
    elif "PermissionError" in str(type(e)):
        return get_friendly_error("E003", detail)
    
    return get_friendly_error(code, detail)


def is_debug_mode() -> bool:
    """检查是否为调试模式"""
    return os.getenv("DEBUG", "").lower() in ["1", "true", "yes"]


# tools 层（延迟加载，避免 db import 冲突）
# store_matcher 在 _call_match_store() 里动态导入，不在模块级别

def _call_match_store(store_name: str, customer_company: str = None,
                      db_config: dict = None,
                      phone: str = None, address: str = None,
                      contact_person: str = None) -> Optional[dict]:
    """动态导入并调用 tools.store_matcher.match_store"""
    from skills.skill_order_to_huading_template.tools._store_matcher import match_store
    return match_store(
        store_name=store_name,
        customer_company=customer_company,
        db_config=db_config,
        phone=phone,
        address=address,
        contact_person=contact_person,
    )

def _is_sku_code(s: str) -> bool:
    """判断字符串是否像商品编码（如 A001, SK241228000106）"""
    return bool(re.match(r'^[A-Z][A-Z0-9]{2,}$', s))


def _parse_item_row(line: str) -> Optional[Dict]:
    """
    解析订单商品行，支持两种格式：
    1. 简单格式：'1. 潮迹潮汕牛肉丸 5 箱'
    2. 多列空格分隔：'1  A001  潮迹潮汕牛肉丸  1×20袋  5  箱'

    Returns: {product_name, spec, quantity, unit, remark, product_code} or None
    """


# ========== 文件下载 URL 生成 ==========
# AWS OpenClaw 配置
AWS_PUBLIC_IP = "13.212.17.85"
AWS_FILE_PORT = 18790


def _get_download_url(output_file: str) -> str:
    """
    生成文件下载 URL
    
    基于 AWS 公网 IP 和文件路径生成下载链接
    格式: http://{IP}:{port}/{filename}
    """
    if not output_file:
        return ""
    
    filename = os.path.basename(output_file)
    # URL 编码中文字符
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    return f"http://{AWS_PUBLIC_IP}:{AWS_FILE_PORT}/{encoded_filename}"
    parts = line.split()
    if len(parts) < 2:
        return None

    n = len(parts)

    # ===== 从右往左解析：数量 + 单位 =====
    i = n - 1
    unit = "件"
    quantity = 1

    while i >= 0:
        p = parts[i]
        if re.match(r'^\d+$', p):
            quantity = int(p)
            i -= 1
            if i >= 0 and parts[i] in ["箱", "件", "袋", "包", "个", "瓶", "桶", "条", "盒", "台"]:
                unit = parts[i]
                i -= 1
            break
        elif p in ["箱", "件", "袋", "包", "个", "瓶", "桶", "条", "盒", "台"]:
            unit = p
            i -= 1
            continue
        else:
            i -= 1
            continue

    # ===== 解析商品名 =====
    remaining = parts[1:i+1]
    has_sku_code = len(remaining) > 0 and _is_sku_code(remaining[0])

    if has_sku_code:
        product_name = remaining[1] if len(remaining) > 1 else remaining[0]
        spec = " ".join(remaining[2:]) if len(remaining) > 2 else ""
    else:
        product_name = remaining[0] if remaining else ""
        spec = " ".join(remaining[1:]) if len(remaining) > 1 else ""

    # ===== 简单正则兜底（处理 '1. 商品名 5 箱' 等格式）=====
    if not product_name or len(product_name) <= 1:
        item_match = re.match(r'^[\-\*\d]+[\.、\s]+(.+?)(?:\s+(\d+)\s*[\u4e00-\u9fa5件个箱袋台]?)?$', line)
        if item_match:
            return {
                "product_name": item_match.group(1).strip(),
                "spec": "",
                "quantity": int(item_match.group(2)) if item_match.group(2) else 1,
                "unit": "件",
                "remark": "",
                "product_code": ""
            }

    return {
        "product_name": product_name,
        "spec": spec,
        "quantity": quantity,
        "unit": unit,
        "remark": "",
        "product_code": ""
    }


class OrderToHuadingTemplate:
    """订单转华鼎出库单模板Skill"""
    
    VERSION = "5.0"  # LLM解析版本
    
    # ========== AI调用约束（方案1：技术层面）==========
    # AI 只能调用这些公开接口，不得直接调用内部工具函数
    __公开接口__ = ['execute', 'tools_parse', 'tools_transform']
    # 内部初始化方法（在 __init__ 期间允许调用）
    __内部初始化__ = ['_load_warehouse_mapping', '_load_field_mapping', '_check_db_connection', '_init_db_repos']
    
    # ========== 必填配置项 ==========
    REQUIRED_CONFIG = {
        "db_config": {
            "required": True,
            "description": "数据库连接配置",
            "fields": {
                "host": {"type": "string", "default": "localhost", "description": "数据库主机地址"},
                "port": {"type": "int", "default": "5432", "description": "数据库端口"},
                "database": {"type": "string", "default": "neo", "description": "数据库名称"},
                "user": {"type": "string", "default": "jinqianfei", "description": "数据库用户名"},
                "password": {"type": "string", "default": "", "description": "数据库密码（必填）"}
            }
        }
    }
    
    # ========== 配置检查 ==========
    @classmethod
    def check_config(cls, db_config: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        检查Skill配置是否完整，返回配置状态和缺失项
        
        Returns:
            Dict: {
                "ready": bool,           # 是否可以运行
                "missing": [field, ...], # 缺失的配置项
                "config_template": {},   # 完整配置模板
                "setup_guide": string    # 配置指引
            }
        """
        missing = []
        
        # 检查 db_config
        if not db_config:
            missing.append("db_config")
        else:
            if not db_config.get("password"):
                missing.append("db_config.password")
        
        setup_guide = """
=== Skill 配置指引 ===

【订单转华鼎出库单模板】需要以下配置：

1️⃣ 数据库配置（db_config）- 必填

  方式一：直接传入 db_config 参数
    skill = OrderToHuadingTemplate(
        db_config={
            "host": "localhost",
            "port": 5432,
            "database": "neo",
            "user": "your_username",
            "password": "your_password"
        }
    )

  方式二：通过环境变量
    export DB_HOST=localhost
    export DB_PORT=5432
    export DB_NAME=neo
    export DB_USER=your_username
    export DB_PASSWORD=your_password

  方式三：通过 .env 文件（项目根目录）
    DB_HOST=localhost
    DB_PORT=5432
    DB_NAME=neo
    DB_USER=your_username
    DB_PASSWORD=your_password

【配置模板】
    db_config = {
        "host": "localhost",     # 数据库主机
        "port": 5432,            # 数据库端口
        "database": "neo",        # 数据库名称（不是 neondb）
        "user": "jinqianfei",    # 数据库用户名
        "password": "***"        # 数据库密码（必填）
    }

=== 配置完成后再使用 ===
"""
        
        config_template = {
            "db_config": {
                "host": os.getenv("DB_HOST", "localhost"),
                "port": int(os.getenv("DB_PORT", "5432")),
                "database": os.getenv("DB_NAME", "neo"),
                "user": os.getenv("DB_USER", "jinqianfei"),
                "password": os.getenv("DB_PASSWORD", "") or "（未配置）"
            }
        }
        
        return {
            "ready": len(missing) == 0,
            "missing": missing,
            "config_template": config_template,
            "setup_guide": setup_guide.strip()
        }
    
    def __getattribute__(self, name: str):
        """拦截内部工具函数调用，防止AI跳过主入口"""
        # 允许所有 dunder 方法和属性
        if name.startswith('__'):
            return object.__getattribute__(self, name)
        
        # 检查是否是内部工具函数
        if name.startswith('_'):
            # 获取公开接口列表
            try:
                公开接口 = object.__getattribute__(self, '__公开接口__')
            except AttributeError:
                公开接口 = []
            
            # 检查是否是允许的内部初始化方法（仅在对象创建期间）
            try:
                内部初始化 = object.__getattribute__(self, '__内部初始化__')
            except AttributeError:
                内部初始化 = []
            
            if name in 内部初始化:
                # 允许内部初始化方法直接调用
                return object.__getattribute__(self, name)
            elif name not in 公开接口:
                # 非公开接口的内部方法，禁止 AI 直接调用
                raise OrderSkillError(
                    code="E001",
                    message=f"禁止直接调用内部函数 '{name}'，请通过 execute() 主入口调用",
                    detail=f"'{name}' 是内部函数，AI 不得直接调用。正确方式：skill.execute(order_input=...)"
                )
        
        return object.__getattribute__(self, name)
    # ========== AI调用约束结束 ==========
    
    def __init__(self, db_config: Dict[str, Any] = None, output_dir: Optional[str] = None):
        """
        初始化Skill配置
        
        Args:
            db_config: 数据库连接配置（必填）
                       支持环境变量读取：
                       - DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD
            output_dir: 输出目录（可选，默认./output）
        """
        # 如果db_config为空，尝试从环境变量读取
        if not db_config:
            db_config = {
                "host": os.getenv("DB_HOST", "localhost"),
                "port": int(os.getenv("DB_PORT", "5432")),
                "database": os.getenv("DB_NAME", "neo"),
                "user": os.getenv("DB_USER", "jinqianfei"),
                "password": os.getenv("DB_PASSWORD", "")
            }
            
            # 如果密码为空，尝试读取 .env 文件
            if not db_config.get("password"):
                env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", ".env")
                if os.path.exists(env_path):
                    try:
                        with open(env_path) as f:
                            for line in f:
                                line = line.strip()
                                if not line or line.startswith("#"):
                                    continue
                                if "=" in line:
                                    k, v = line.split("=", 1)
                                    k = k.strip()
                                    if k == "DB_PASSWORD":
                                        db_config["password"] = v.strip()
                                        break
                    except Exception:
                        pass
        
        # 检查配置是否完整
        config_status = self.check_config(db_config)
        if not config_status["ready"]:
            missing = config_status["missing"]
            guide = config_status["setup_guide"]
            raise OrderSkillError(
                code="E401",
                message=f"Skill配置不完整，缺少：{', '.join(missing)}",
                detail=guide
            )

        # 移除 SOCKS 代理，避免 LLM 调用时出错（socksio 未安装）
        for k in ["SOCKS_PROXY", "ALL_PROXY", "HTTPS_PROXY", "HTTP_PROXY"]:
            os.environ.pop(k, None)
        os.environ.pop("no_proxy", None)
        os.environ.pop("NO_PROXY", None)

        # 从 .env 文件加载环境变量（如果存在）
        env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", ".env")
        if os.path.exists(env_path):
            try:
                with open(env_path) as f:
                    for line in f:
                        line = line.strip()
                        if not line or line.startswith("#"):
                            continue
                        if "=" in line:
                            k, v = line.split("=", 1)
                            if k not in os.environ:
                                os.environ[k.strip()] = v.strip()
            except Exception:
                pass

        self.shipper_id = None  # 不再直接传入，通过门店匹配获取
        self.db_config = db_config
        self.output_dir = output_dir or "./output"
        
        # 从数据库加载仓库编码映射
        self.warehouse_code_map = object.__getattribute__(self, '_load_warehouse_mapping')()
        
        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)

    
    # LLM 解析提示词模板
    LLM_PARSE_PROMPT = """你是一个订单数据提取专家。请从客户订单中提取结构化信息。

【重要】一个订单可能包含多个门店，每个门店有独立的商品明细。请将所有门店都提取出来。

【输出格式】
请直接输出JSON格式（不要有其他文字）：
{{
  "confidence": 0.95,
  "stores": {{
    "门店A名称": {{
      "store_name": "门店A完整名称",
      "contact_person": "联系人姓名",
      "phone": "联系电话",
      "address": "收货地址",
      "order_no": "订单编号（如有）",
      "items": [
        {{
          "seq": 1,
          "product_name": "商品名称",
          "spec": "规格（如有）",
          "quantity": 数量,
          "unit": "单位",
          "product_code": "商品编码（如有）",
          "remark": "商品备注（如有）"
        }}
      ]
    }},
    "门店B名称": {{
      "store_name": "门店B完整名称",
      "contact_person": "...",
      "phone": "...",
      "address": "...",
      "items": [...]
    }}
  }},
  "warnings": ["告警信息"]
}}

【重要规则】
1. 一个订单文件如果包含多个门店，必须把所有门店都列在 stores 对象里
2. 每个门店的 items 只包含该门店的商品
3. 如果订单只有一个门店，仍需放在 stores 对象里（单键）
4. product_name 必须是完整商品名称，quantity 必须是数字
5. 只返回JSON，不要有其他解释文字

【订单内容】：
{content}
"""


    FIELD_ALIAS_MAPPING = {
        # 门店相关
        "store_name": ["门店", "店名", "店铺名称", "店铺", "shop", "shop_name", "name", "收货方", "客户名称", "公司名称", "客户"],
        
        # 联系方式
        "phone": ["电话", "手机", "联系电话", "tel", "mobile", "号码", "联系手机", "手机号"],
        "contact_person": ["联系人", "收货人", "contact", "收货人姓名", "收件人"],
        "address": ["地址", "收货地址", "配送地址", "addr", "详细地址"],
        
        # 商品
        "product_name": ["商品", "货品", "产品名", "product", "goods", "品名", "商品名称", "名称", "货品名称"],
        "quantity": ["数量", "件数", "箱数", "qty", "num", "订单数量", "订货数量"],
        "unit": ["单位", "件", "箱", "unit", "包装", "包装单位", "计量单位"],
        "spec": ["规格", "spec", "规格型号", "型号", "商品规格"],
        "product_code": ["编码", "商品编码", "SKU", "货号", "编号"],
        "remark": ["备注", "note", "备注信息"],
        
        # 订单
        "items": ["items", "商品明细", "商品列表", "订货明细", "明细"],
        
        # 订单
        "order_no": ["订单号", "单号", "订单编号", "order", "order_no", "送货单号", "单据编号"],
        
        # 货主公司
        "customer_company": ["客户公司", "公司名", "货主公司", "供应商", "发货方", "厂商"],
    }
    
    # 标准字段名列表（用于校验）
    STANDARD_FIELDS = list(FIELD_ALIAS_MAPPING.keys())
    
    # 华鼎模板31字段
    HUADING_FIELDS = [
        "序号", "门店编号", "门店三方编码", "仓库编码", "加急程度\n（0：普通，1：加急）",
        "商品SKU编号", "商品三方SPEC编号", "单位类型", "出库数量",
        "指定库存状态", "出库类型", "配送方式", "指定车型（专配）",
        "是否垫付", "付款方式", "快递公司", "单价", "总金额",
        "是否制定批次", "批次号", "生产日期", "备注", "生产厂家编号",
        "门店收货地址编码", "三方单号", "业务模式", "业务类型",
        "收货人", "联系电话", "收货地址", "C端快递公司"
    ]
    
    # 有默认值的字段
    DEFAULT_VALUES = {
        "加急程度": 0,
        "指定库存状态": "正常",
        "出库类型": 201,
        "配送方式": "共配",
        "是否垫付": "否",
        "是否制定批次": "否"
    }
    
    # ========== 字段名标准化相关方法 ==========
    
    @staticmethod
    def find_standard_field(input_name: str) -> Optional[str]:
        """
        将任意字段名映射到标准字段名
        
        Args:
            input_name: 原始字段名
            
        Returns:
            标准字段名，或None如果未找到
        """
        if not input_name:
            return None
        
        input_name = str(input_name).strip().lower()
        
        for std_field, aliases in OrderToHuadingTemplate.FIELD_ALIAS_MAPPING.items():
            # 完全匹配（不区分大小写）
            if input_name == std_field.lower():
                return std_field
            
            # 别名匹配
            for alias in aliases:
                if input_name == alias.lower():
                    return std_field
        
        return None
    
    def normalize_ai_result(self, raw_data: Dict) -> Tuple[Dict, List[str]]:
        """
        AI返回的字段名标准化
        
        Args:
            raw_data: AI返回的原始数据 dict
            
        Returns:
            (标准化后的数据, 警告信息列表)
        """
        normalized = {}
        warnings = []
        unknown_fields = []
        
        for ai_field_name, value in raw_data.items():
            # 跳过特殊字段
            if ai_field_name.startswith("__") or ai_field_name in ["raw_text"]:
                continue
            
            std_field = self.find_standard_field(ai_field_name)
            
            if std_field:
                normalized[std_field] = value
                if std_field != ai_field_name and ai_field_name.lower() != std_field.lower():
                    warnings.append(f"字段名'{ai_field_name}'→'{std_field}'")
            else:
                unknown_fields.append(ai_field_name)
        
        if unknown_fields:
            warnings.append(f"未知字段名{len(unknown_fields)}个，已忽略: {', '.join(unknown_fields[:5])}")
        
        return normalized, warnings
    
    def normalize_excel_headers(self, df_raw, header_row_hints: List[int] = None) -> Tuple[Dict[int, str], List[str]]:
        """
        Excel按表头名读取，找到表头行并建立列索引→标准字段名的映射
        
        Args:
            df_raw: pandas DataFrame
            header_row_hints: 可能的表头行索引列表，默认[3, 4, 5, 6]
            
        Returns:
            (列索引→标准字段名映射, 警告信息列表)
        """
        warnings = []
        header_row_hints = header_row_hints or [3, 4, 5, 6]
        
        col_mapping = {}  # {列索引: 标准字段名}
        
        for row_idx in header_row_hints:
            if row_idx >= len(df_raw):
                continue
            
            row = df_raw.iloc[row_idx]
            temp_mapping = {}
            matched_count = 0
            
            for col_idx, cell in enumerate(row):
                if cell is None or (hasattr(cell, '__iter__') and not isinstance(cell, str)):
                    if pd.isna(cell) if hasattr(pd, 'isna') else not cell:
                        continue
                
                cell_str = str(cell).strip() if cell else ""
                if not cell_str:
                    continue
                
                std_field = self.find_standard_field(cell_str)
                if std_field:
                    temp_mapping[col_idx] = std_field
                    matched_count += 1
            
            # 选择匹配数最多的行作为表头行
            if matched_count > len(col_mapping):
                col_mapping = temp_mapping
                if matched_count > 0:
                    warnings.append(f"表头行推断: 第{row_idx + 1}行，匹配{len(col_mapping)}个字段")
            
            # 如果已经匹配到关键字段（store_name和items相关），可以提前结束
            if all(v in col_mapping.values() for v in ["store_name", "product_name", "quantity"]):
                break
        
        if not col_mapping:
            warnings.append("未能识别Excel表头行，将使用默认列映射")
        
        return col_mapping, warnings
    
    def normalize_text_fields(self, text: str) -> Tuple[Dict, List[str]]:
        """
        文本按字段名映射表匹配
        支持多种格式：字段名：值、字段名 值
        
        Args:
            text: 原始文本
            
        Returns:
            (字段→值映射, 警告信息列表)
        """
        normalized = {}
        warnings = []
        
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        
        for line in lines:
            # 跳过空行
            if not line:
                continue
            
            matched = False
            
            for std_field, aliases in self.FIELD_ALIAS_MAPPING.items():
                # 跳过items的处理
                if std_field == "items":
                    continue
                
                for alias in aliases:
                    # 检查字段名是否在行首
                    if line.startswith(alias):
                        # 截取字段名之后的内容
                        value = line[len(alias):]
                        # 去除开头的标点符号和空格（：:,，、等）
                        value = re.sub(r'^[\s：:,，、]+', '', value)
                        
                        if std_field not in normalized and value:
                            normalized[std_field] = value
                            warnings.append(f"文本字段'{alias}'→'{std_field}'")
                            matched = True
                        
                        if matched:
                            break
                if matched:
                    break
        
        return normalized, warnings
    
    def validate_normalized_data(self, data: Dict) -> Dict[str, Any]:
        """
        校验标准化后的数据
        
        Args:
            data: 标准化后的数据 dict
            
        Returns:
            {
                "is_valid": bool,
                "warnings": List[str],
                "errors": List[str],
                "suggestions": List[str]
            }
        """
        result = {
            "is_valid": True,
            "warnings": [],
            "errors": [],
            "suggestions": []
        }
        
        # 1. 检查必要字段
        if not data.get("store_name"):
            result["errors"].append("缺少门店名字段")
            result["is_valid"] = False
        
        if not data.get("items") or len(data.get("items", [])) == 0:
            # 检查items是否是dict列表或需要进一步处理
            items = data.get("items", [])
            if isinstance(items, list) and len(items) == 0:
                result["warnings"].append("商品明细为空")
            elif not isinstance(items, list):
                result["errors"].append("商品明细格式错误")
                result["is_valid"] = False
        
        # 2. 检查值格式
        phone = data.get("phone", "")
        if phone:
            # 简单手机号格式校验（中国大陆）
            phone_clean = re.sub(r'\\D', '', phone)
            if len(phone_clean) > 0 and (len(phone_clean) < 7 or len(phone_clean) > 13):
                result["warnings"].append(f"电话号码格式可疑: {phone}")
        
        # 3. 检查数量
        items = data.get("items", [])
        if isinstance(items, list):
            for idx, item in enumerate(items):
                if isinstance(item, dict):
                    qty = item.get("quantity", 0)
                    if qty and (not isinstance(qty, (int, float)) or qty < 0):
                        result["errors"].append(f"商品{idx + 1}数量异常: {qty}")
                        result["is_valid"] = False
        
        return result
    
    def _format_success_message(self, store_names: str, item_count: int,
                                   matched_count: int, unmatched_count: int,
                                   file_name: str, has_issues: bool = False) -> str:
        """
        格式化成功返回的用户消息（支持飞书等多平台展示）
        
        Args:
            store_names: 门店名称列表
            item_count: 总商品数
            matched_count: 匹配成功数
            unmatched_count: 未匹配数
            file_name: 输出文件名
            has_issues: 是否有问题需要关注
        
        Returns:
            str: 友好的用户消息
        """

        # 基础信息
        lines = []
        lines.append(f"✅ 订单处理完成")
        lines.append(f"")
        lines.append(f"📦 订单信息")
        lines.append(f"   门店：{store_names}")
        lines.append(f"   商品数：{item_count}条")
        lines.append(f"")
        lines.append(f"📊 匹配结果")
        lines.append(f"   ✅ 匹配成功：{matched_count}条")
        
        if unmatched_count > 0:
            lines.append(f"   ⚠️ 未匹配：{unmatched_count}条（需人工确认）")
        else:
            lines.append(f"   ⬜ 未匹配：0条")
        
        lines.append(f"")
        lines.append(f"📁 输出文件")
        lines.append(f"   {file_name}")
        lines.append(f"")
        
        if has_issues:
            lines.append(f"⚠️ 温馨提示：有{unmatched_count}条商品未匹配成功，请检查后确认")
        else:
            lines.append(f"💡 请下载文件查看详情，确认无误后提交审核")
        
        return "\n".join(lines)
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        获取文件信息（用于发送文件给用户）
        
        Args:
            file_path: 文件路径
        
        Returns:
            Dict: {exists, file_name, file_size, mime_type, path}
        """
        if not os.path.exists(file_path):
            return {
                "exists": False,
                "file_name": os.path.basename(file_path),
                "error": "文件不存在"
            }
        
        file_size = os.path.getsize(file_path)
        # 转换为KB或MB
        if file_size > 1024 * 1024:
            size_str = f"{file_size / (1024 * 1024):.1f} MB"
        elif file_size > 1024:
            size_str = f"{file_size / 1024:.0f} KB"
        else:
            size_str = f"{file_size} B"
        
        return {
            "exists": True,
            "file_name": os.path.basename(file_path),
            "file_size": file_size,
            "file_size_str": size_str,
            "path": file_path,
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    
    def prepare_file_for_send(self, file_path: str) -> Dict[str, Any]:
        """
        准备文件发送所需信息（供OpenClaw/飞书等平台使用）
        
        Args:
            file_path: 文件路径
        
        Returns:
            Dict: 包含文件信息和发送所需的元数据
        """
        file_info = self.get_file_info(file_path)
        
        if not file_info["exists"]:
            return {
                "success": False,
                "error": file_info["error"]
            }
        
        return {
            "success": True,
            "file_path": file_path,
            "file_name": file_info["file_name"],
            "file_size_str": file_info["file_size_str"],
            "can_send": True,
            # OpenClaw message工具需要的字段
            "for_message": {
                "media": file_path,  # 支持文件路径
                "file_name": file_info["file_name"]
            }
        }

    
    def _load_warehouse_mapping(self) -> Dict[str, str]:
        """从数据库加载仓库编码映射"""
        try:
            import psycopg2
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            cur.execute("SELECT warehouse_name, warehouse_code FROM warehouse_code_mapping")
            rows = cur.fetchall()
            
            conn.close()
            
            mapping = {}
            for name, code in rows:
                mapping[name] = code
            
            return mapping
            
        except Exception as e:
            print(f"加载仓库编码映射失败: {e}")
            return {}
    
    def _get_warehouse_code(self, warehouse_name: str) -> str:
        """获取仓库编码，找不到则报错"""
        if not warehouse_name:
            raise ValueError(f"门店的仓库名称为空，请检查门店数据")
        
        warehouse_code = self.warehouse_code_map.get(warehouse_name)
        if not warehouse_code:
            available = ", ".join(self.warehouse_code_map.keys())
            raise ValueError(
                f"仓库'{warehouse_name}'在仓库编码映射表中未找到！\n"
                f"请检查仓库编码表，确保该仓库已配置。\n"
                f"当前可用的仓库：{available}"
            )
        
        return warehouse_code
    
    def _detect_input_type(self, order_input: str) -> str:
        """
        自动检测输入类型
        
        Returns:
            "excel" / "image" / "pdf" / "text"
        """
        if not order_input:
            return "text"
        
        # 1. 检查是否为文件路径
        if os.path.exists(order_input):
            ext = os.path.splitext(order_input)[1].lower()
            if ext in [".xlsx", ".xls"]:
                return "excel"
            elif ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif"]:
                return "image"
            elif ext == ".pdf":
                return "pdf"
            else:
                # 未知扩展名，尝试作为图片处理
                return "image"
        
        # 2. 检查扩展名字符串
        input_str = str(order_input).strip().lower()
        if input_str.endswith(".xlsx") or input_str.endswith(".xls"):
            return "excel"
        elif input_str.endswith(".jpg") or input_str.endswith(".png") or input_str.endswith(".jpeg"):
            return "image"
        elif input_str.endswith(".pdf"):
            return "pdf"
        
        # 3. 纯文本判断（包含换行或结构化标记）
        if "\n" in order_input or "订单号" in order_input or "门店" in order_input:
            return "text"
        
        # 4. 默认当作图片路径处理（网络URL或相对路径）
        return "image"
    
    def _parse_image(self, image_path: str) -> Optional[Dict]:
        """
        图片解析 - 返回需要外部OCR的标记
        
        由于图像识别需要AI能力，此方法返回特殊状态码 "NEED_OCR"
        由调用方使用image工具进行识别后，将结果传入_handle_ocr_result()
        
        Args:
            image_path: 图片文件路径
            
        Returns:
            {"__need_ocr__: True, "image_path": image_path} 或 None
        """
        if not image_path or not os.path.exists(image_path):
            return None
        
        # 返回需要OCR的标记
        return {"__need_ocr__": True, "image_path": image_path}
    
    def _handle_ocr_result(self, ocr_result: Dict) -> Optional[Dict]:
        """
        处理OCR识别结果，将结构化数据规范化
        
        Args:
            ocr_result: OCR工具返回的结构化数据
            
        Returns:
            规范化后的订单数据字典
        """
        if not ocr_result or not isinstance(ocr_result, dict):
            return None
        
        # 如果OCR结果已经是结构化数据，直接规范化
        return object.__getattribute__(self, '_normalize_extracted_data')(ocr_result)
    
    def _parse_pdf(self, pdf_path: str) -> Optional[Dict]:
        """
        PDF解析 - 返回需要外部OCR的标记
        
        由于PDF识别需要AI能力，此方法返回特殊状态码 "__need_pdf_ocr__"
        由调用方使用pdf工具进行识别后，将结果传入_handle_pdf_ocr_result()
        
        Args:
            pdf_path: PDF文件路径
            
        Returns:
            {"__need_pdf_ocr__": True, "pdf_path": pdf_path} 或 None
        """
        if not pdf_path or not os.path.exists(pdf_path):
            return None
        
        # 返回需要PDF OCR的标记
        return {"__need_pdf_ocr__": True, "pdf_path": pdf_path}
    
    def _handle_pdf_ocr_result(self, pdf_ocr_result: Dict) -> Optional[Dict]:
        """
        处理PDF OCR识别结果，将结构化数据规范化
        
        Args:
            pdf_ocr_result: PDF工具返回的结构化数据
            
        Returns:
            规范化后的订单数据字典
        """
        if not pdf_ocr_result or not isinstance(pdf_ocr_result, dict):
            return None
        
        # 如果OCR结果已经是结构化数据，直接规范化
        return object.__getattribute__(self, '_normalize_extracted_data')(pdf_ocr_result)
    
    def _parse_text(self, text: str) -> Optional[Dict]:
        """
        从纯文本/粘贴内容中解析订单数据
        支持多种常见格式
        """
        return object.__getattribute__(self, '_normalize_extracted_data')({"raw_text": text})
    
    def _normalize_extracted_data(self, extracted: Dict) -> Optional[Dict]:
        """
        规范化从OCR/文本提取的数据，适配标准订单格式
        先进行字段名标准化，再处理数据
        """
        raw_text = extracted.get("raw_text", "")
        
        # 如果是纯文本，先解析
        if raw_text:
            return object.__getattribute__(self, '_parse_raw_text')(raw_text)
        
        # ========== Step 1: 字段名标准化 ==========
        normalized_fields, norm_warnings = self.normalize_ai_result(extracted)
        
        # 如果标准化后没有store_name，可能是识别问题
        if not normalized_fields.get("store_name") and not extracted.get("items"):
            # 尝试原始数据的keys
            if not normalized_fields:
                normalized_fields = {k: v for k, v in extracted.items() 
                                   if not k.startswith("__") and k != "raw_text"}
        
        # ========== Step 2: 提取各字段 ==========
        order_no = str(normalized_fields.get("order_no", extracted.get("order_no", ""))).strip()
        store_name = str(normalized_fields.get("store_name", extracted.get("store_name", ""))).strip()
        
        # 清理门店名称
        for prefix in ['河北-', '天津-', '沧州-']:
            if store_name.startswith(prefix):
                store_name = store_name[len(prefix):]
                break
        
        # 提取商品明细
        items_raw = normalized_fields.get("items", extracted.get("items", []))
        items = []
        for idx, item in enumerate(items_raw if isinstance(items_raw, list) else [], start=1):
            if not isinstance(item, dict):
                continue
            
            # 支持商品名字段名变体
            product_name = None
            for pn_key in ["product_name", "product", "goods", "name", "品名"]:
                if pn_key in item:
                    product_name = str(item.get(pn_key, "")).strip()
                    break
            
            if not product_name or product_name == "nan":
                continue
            
            # 数量处理
            qty = None
            for qty_key in ["quantity", "qty", "num", "数量"]:
                if qty_key in item:
                    qty = item.get(qty_key, 0)
                    break
            if qty is None:
                qty = 0
            if isinstance(qty, str):
                qty = int(float(re.sub(r'[^0-9.]', '', qty)))
            qty = int(qty) if qty else 0
            
            items.append({
                "seq": item.get("seq", idx),
                "product_code": str(item.get("product_code", item.get("code", ""))).strip(),
                "product_name": product_name,
                "spec": str(item.get("spec", item.get("规格", ""))).strip(),
                "quantity": qty,
                "unit": str(item.get("unit", item.get("单位", "件"))).strip(),
                "remark": str(item.get("remark", item.get("note", ""))).strip()
            })
        
        if not store_name and not items:
            return None
        
        # 提取客户公司名称（货主公司，用于优先匹配货主ID）
        customer_company = str(normalized_fields.get("customer_company", extracted.get("customer_company", ""))).strip()
        
        return {
            "order_no": order_no,
            "store_name": store_name,
            "contact_person": str(normalized_fields.get("contact_person", extracted.get("contact_person", ""))).strip(),
            "phone": str(normalized_fields.get("phone", extracted.get("phone", ""))).strip(),
            "address": str(normalized_fields.get("address", extracted.get("address", ""))).strip(),
            "items": items,
            "customer_company": customer_company,
            "_norm_warnings": norm_warnings  # 保留警告信息供后续使用
        }
    

    def tools_parse(self, order_input: str, order_type: str = "auto") -> Dict[str, Any]:
        """
        Step 1: 调用 tools/order_parser.parse() 解析订单
        
        职责: 将任意格式订单解析为原始结构化JSON
        - Excel/图片/PDF/Word/文字 全部经过 LLM 解析
        - 返回原始 JSON（多门店 stores{} 格式）
        
        Args:
            order_input: 文件路径 或 文本内容
            order_type: auto/excel/image/pdf/word/text
        
        Returns:
            order_parser.parse() 的返回结果
        """
        from skills.skill_order_to_huading_template.tools._order_parser import parse
        return parse(order_input, order_type=order_type)


    def tools_transform(self, order_data: Dict) -> Dict[str, Any]:
        """
        Step 2: 调用 tools/field_transformer.transform() 规则库标准化
        
        职责: 将 LLM 原始输出字段名映射为统一字段名
        - 加载客户 YAML 规则库
        - 字段名标准化（LLM输出字段名 → 统一字段名）
        - 值校验与标准化
        
        Args:
            order_data: tools_parse() 返回的原始解析结果
        
        Returns:
            field_transformer.transform() 的返回结果（统一 JSON）
        """
        from skills.skill_order_to_huading_template.tools._field_transformer import transform
        return transform(order_data)

    def parse_with_llm(self, content: str, content_type: str = "text") -> Dict[str, Any]:
        """
        使用 LLM 解析订单内容（文本/Excel/图片）
        支持 MiniMax 和 OpenAI 两种后端
        
        Args:
            content: 订单内容（文本内容，或文件路径）
            content_type: "text" / "excel" / "image"
            
        Returns:
            {
                "success": bool,
                "confidence": float,
                "store_name": str,
                "contact_person": str,
                "phone": str,
                "address": str,
                "order_no": str,
                "shipper_name": str,
                "warehouse_name": str,
                "remark": str,
                "items": [...],
                "warnings": [...],
                "llm_raw_output": str  # 原始LLM输出
            }
        """
        try:
            import json
            import os
            
            # 优先使用 MiniMax
            api_key = os.getenv("MINIMAX_API_KEY") or os.getenv("OPENAI_API_KEY")
            if not api_key:
                raise ValueError("未设置 MINIMAX_API_KEY 或 OPENAI_API_KEY")
            
            # 从 .env 文件加载环境变量（如果存在）
            env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", ".env")
            if os.path.exists(env_path):
                try:
                    with open(env_path) as f:
                        for line in f:
                            line = line.strip()
                            if not line or line.startswith("#"):
                                continue
                            if "=" in line:
                                k, v = line.split("=", 1)
                                if k.strip() not in os.environ:
                                    os.environ[k.strip()] = v.strip()
                except Exception:
                    pass
            
            # 再次确认API key已加载（覆盖之前的检查）
            api_key = os.getenv("MINIMAX_API_KEY") or os.getenv("OPENAI_API_KEY")
            
            # 如果是Excel文件，先读取内容
            if content_type == "excel" and os.path.exists(content):
                import pandas as pd
                df = pd.read_excel(content, header=None)
                # 将Excel转为文本描述
                content = object.__getattribute__(self, '_excel_to_text')(df)
            
            # 构建提示词
            prompt = self.LLM_PARSE_PROMPT.format(content=content[:8000] if len(content) > 8000 else content)
            
            # 判断使用哪个后端
            if os.getenv("MINIMAX_API_KEY"):
                # 使用 MiniMax API
                import openai
                
                client = openai.OpenAI(
                    api_key=api_key,
                    base_url="https://api.minimax.chat/v1"
                )
                
                response = client.chat.completions.create(
                    model="MiniMax-M2.7",
                    messages=[
                        {"role": "system", "content": "你是一个专业的订单数据提取专家。请从用户提供的订单内容中准确提取结构化信息。"},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.1,
                    max_tokens=4000
                )
            else:
                # 使用 OpenAI API
                import openai
                
                client = openai.OpenAI(api_key=api_key)
                
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {"role": "system", "content": "你是一个专业的订单数据提取专家。请从用户提供的订单内容中准确提取结构化信息。"},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.1,
                    max_tokens=4000
                )
            
            raw_output = response.choices[0].message.content.strip()
            
            # 解析 JSON - 健壮的 JSON 提取（搜索所有代码块）
            # MiniMax M2 模型可能输出多轮思考过程和多个代码块
            # 策略：搜索所有 ```...``` 块，尝试解析每个为 JSON
            json_blocks = re.findall(r"```(?:json)?\s*([\s\S]*?)```", raw_output)
            llm_data = None
            last_parse_error = ""
            for block in json_blocks:
                block = block.strip()
                if not block:
                    continue
                try:
                    llm_data = json.loads(block)
                    break  # 成功解析
                except (json.JSONDecodeError, ValueError) as e:
                    last_parse_error = str(e)
                    continue
            
            # 如果所有代码块都解析失败，尝试直接解析（移除思考标签后）
            if llm_data is None:
                # 移除思考标签后再试
                cleaned = re.sub(r'<think>.*?</think>', '', raw_output, flags=re.DOTALL)
                cleaned = re.sub(r'<!--.*?-->', '', cleaned, flags=re.DOTALL)
                cleaned = cleaned.strip()
                try:
                    llm_data = json.loads(cleaned)
                except (json.JSONDecodeError, ValueError):
                    return {
                        "success": False,
                        "error": f"LLM输出无法解析为JSON（代码块解析失败: {last_parse_error}）",
                        "llm_raw_output": raw_output[:500]
                    }


            # ========== 新格式：多门店 (stores 顶层) ==========
            if "stores" in llm_data:
                result = {
                    "success": True,
                    "confidence": llm_data.get("confidence", 0.9),
                    "stores": llm_data.get("stores", {}),  # 保持原始结构供 _normalize_llm_result 处理
                    "order_no": "",
                    "shipper_name": "",
                    "warehouse_name": "",
                    "warnings": llm_data.get("warnings", []),
                    "llm_raw_output": raw_output,
                }
                return result

            # ========== 旧格式：单门店 (store_info + items) ==========
            store_info = llm_data.get("store_info", {})
            order_info = llm_data.get("order_info", {})
            items = llm_data.get("items", [])

            # 清理门店名（去掉前缀）
            store_name = store_info.get("name", "")
            for prefix in ['河北-', '天津-', '沧州-']:
                if store_name.startswith(prefix):
                    store_name = store_name[len(prefix):]
                    break

            result = {
                "success": True,
                "confidence": llm_data.get("confidence", 0.9),
                "store_name": store_name,
                "contact_person": store_info.get("contact_person", ""),
                "phone": store_info.get("phone", ""),
                "address": store_info.get("address", ""),
                "order_no": order_info.get("order_no", ""),
                "shipper_name": order_info.get("shipper_name", ""),
                "warehouse_name": order_info.get("warehouse_name", ""),
                "remark": order_info.get("remark", ""),
                "items": items,
                "warnings": llm_data.get("warnings", []),
                "llm_raw_output": raw_output
            }

            return result
            
        except Exception as e:
            return {
                "success": False,
                "confidence": 0,
                "error": str(e),
                "store_name": "",
                "contact_person": "",
                "phone": "",
                "address": "",
                "order_no": "",
                "shipper_name": "",
                "warehouse_name": "",
                "remark": "",
                "items": [],
                "warnings": [f"LLM解析失败: {str(e)}"],
                "llm_raw_output": ""
            }
    
    def _excel_to_text(self, df) -> str:
        """将 Excel DataFrame 转为文本描述，包含所有数据行"""
        lines = []
        lines.append("【Excel订单内容开始】")
        
        # 获取所有行（不限制行数）
        # 但限制每行列数，避免过于冗长
        for i in range(len(df)):
            row_data = []
            for j in range(min(len(df.columns), 10)):
                val = df.iloc[i, j]
                if val is not None and str(val) != 'nan' and str(val).strip():
                    row_data.append(str(val).strip())
            if row_data:
                lines.append(f"行{i+1}: {' | '.join(row_data)}")
        
        lines.append("【Excel订单内容结束】")
        return "\n".join(lines)
    
    def _normalize_llm_result(self, llm_result: Dict) -> Dict:
        """
        将 LLM 解析结果转换为标准订单数据格式

        支持两种格式：
        - 新格式（多门店）：stores = {门店名: {store_name, items, ...}}
        - 旧格式（单门店）：store_info + items
        """
        # ========== 新格式：多门店 ==========
        stores_data = llm_result.get("stores", {})
        if stores_data:
            # 合并所有门店的 items（用于整体展示）
            all_items = []
            for store_key, store_data in stores_data.items():
                for idx, item in enumerate(store_data.get("items", []), start=1):
                    all_items.append({
                        "seq": item.get("seq", idx),
                        "product_code": str(item.get("product_code", "")).strip(),
                        "product_name": str(item.get("product_name", "")).strip(),
                        "spec": str(item.get("spec", "")).strip(),
                        "quantity": int(item.get("quantity", 0)),
                        "unit": str(item.get("unit", "件")).strip(),
                        "remark": str(item.get("remark", "")).strip(),
                        "_store_name": store_data.get("store_name", store_key),  # 标记来源门店
                    })
            return {
                "order_no": llm_result.get("order_no", ""),
                "store_name": "",  # 多门店时不确定
                "stores": stores_data,  # 保留原始多门店结构
                "items": all_items,  # 合并所有门店商品
                "_confidence": llm_result.get("confidence", 0.9),
                "_llm_warnings": llm_result.get("warnings", []),
                "_multi_store": True,
            }

        # ========== 旧格式：单门店（兼容）==========
        items = []
        for idx, item in enumerate(llm_result.get("items", []), start=1):
            items.append({
                "seq": item.get("seq", idx),
                "product_code": str(item.get("product_code", "")).strip(),
                "product_name": str(item.get("product_name", "")).strip(),
                "spec": str(item.get("spec", "")).strip(),
                "quantity": int(item.get("quantity", 0)),
                "unit": str(item.get("unit", "件")).strip(),
                "remark": str(item.get("remark", "")).strip()
            })

        return {
            "order_no": llm_result.get("order_no", ""),
            "store_name": llm_result.get("store_name", ""),
            "contact_person": llm_result.get("contact_person", ""),
            "phone": llm_result.get("phone", ""),
            "address": llm_result.get("address", ""),
            "shipper_name": llm_result.get("shipper_name", ""),
            "warehouse_name": llm_result.get("warehouse_name", ""),
            "remark": llm_result.get("remark", ""),
            "items": items,
            "_confidence": llm_result.get("confidence", 0.9),
            "_llm_warnings": llm_result.get("warnings", []),
            "_llm_raw_output": llm_result.get("llm_raw_output", ""),
            "_multi_store": False,
        }
    
    def _parse_raw_text(self, text: str) -> Optional[Dict]:
        """
        解析原始粘贴文本，使用字段名映射表匹配
        支持多种客户格式的字段名
        
        匹配方式：
        - 遍历每一行
        - 用字段名映射表查找这一行是否包含某个字段名
        - 如果包含，截取字段名后面的内容作为值（自动去除标点符号和空格）
        """
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        warnings = []
        
        order_no = ""
        store_name = ""
        contact_person = ""
        phone = ""
        address = ""
        items = []
        
        # 用于记录哪些字段已匹配
        matched_fields = set()
        
        for line in lines:
            # 跳过空行和注释
            if not line or line.startswith("#"):
                continue
            
            # 检查这一行是否是字段行
            line_matched = False
            
            for std_field, aliases in self.FIELD_ALIAS_MAPPING.items():
                # 跳过items的处理（items是列表，需要特殊处理）
                if std_field == "items":
                    continue
                    
                for alias in aliases:
                    # 查找字段名在行中的位置（行首）
                    if line.startswith(alias):
                        # 截取字段名之后的内容
                        value = line[len(alias):]
                        # 去除开头的标点符号和空格（：:,，、等）
                        value = re.sub(r'^[\s：:,，、]+', '', value)
                        
                        if std_field == "store_name" and not store_name:
                            store_name = value
                            matched_fields.add("store_name")
                            warnings.append(f"文本字段'{alias}'→'store_name'")
                            line_matched = True
                        elif std_field == "order_no" and not order_no:
                            order_no = value
                            matched_fields.add("order_no")
                            line_matched = True
                        elif std_field == "contact_person" and not contact_person:
                            contact_person = value
                            matched_fields.add("contact_person")
                            line_matched = True
                        elif std_field == "phone" and not phone:
                            phone = value
                            matched_fields.add("phone")
                            line_matched = True
                        elif std_field == "address" and not address:
                            address = value
                            matched_fields.add("address")
                            line_matched = True
                        
                        if line_matched:
                            break
                if line_matched:
                    break
            
            if line_matched:
                continue
            
            # ========== 商品明细识别 ==========
            # 商品明细识别（支持简单格式 和 多列空格分隔格式）
            item_result = _parse_item_row(line)
            if item_result:
                items.append({
                    "seq": len(items) + 1,
                    "product_code": item_result.get("product_code", ""),
                    "product_name": item_result["product_name"],
                    "spec": item_result.get("spec", ""),
                    "quantity": item_result["quantity"],
                    "unit": item_result.get("unit", "件"),
                    "remark": item_result.get("remark", "")
                })
                continue
            
            # 如果这行只有中文/英文/数字，且较长，可能是门店名
            if re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9]{4,}$', line):
                if not store_name and "store_name" not in matched_fields:
                    store_name = line
                    warnings.append(f"推断门店名: '{line}'")
        
        # 清理门店名
        for prefix in ['河北-', '天津-', '沧州-']:
            if store_name.startswith(prefix):
                store_name = store_name[len(prefix):]
                break
        
        if not store_name and not items:
            warnings.append("未能从文本中提取到门店名和商品明细")
        
        return {
            "order_no": order_no,
            "store_name": store_name,
            "contact_person": contact_person,
            "phone": phone,
            "address": address,
            "items": items,
            "_norm_warnings": warnings
        }
        
        # 如果这行只有中文/英文/数字，且较长，可能是门店名
        if re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9]{4,}$', line):
            if not store_name and "store_name" not in matched_fields:
                store_name = line
                warnings.append(f"推断门店名: '{line}'")
        
        # 清理门店名
        for prefix in ['河北-', '天津-', '沧州-']:
            if store_name.startswith(prefix):
                store_name = store_name[len(prefix):]
                break
        
        if not store_name and not items:
            warnings.append("未能从文本中提取到门店名和商品明细")
        
        return {
            "order_no": order_no,
            "store_name": store_name,
            "contact_person": contact_person,
            "phone": phone,
            "address": address,
            "items": items,
            "_norm_warnings": warnings
        }
    
    def execute(self, order_input: str = None, output_file: str = None, order_type: str = "auto", 
                ocr_result: Dict = None, confirmed_store: Dict = None, 
                order_data_cache: Dict = None) -> Dict[str, Any]:
        """
        执行订单转华鼎模板（支持多格式输入）
        
        Args:
            order_input: 订单输入（文件路径/图片路径/文字内容）
            output_file: 输出文件路径
            order_type: 输入类型（"auto"/"excel"/"image"/"pdf"/"text"）
            ocr_result: 图片/PDF OCR识别结果（当图片/PDF解析返回NEED_OCR时，由调用方识别后传入）
        
        Returns:
            Dict包含:
            - success: 是否成功
            - need_ocr: 是否需要OCR识别（当图片无法直接处理时返回True）
            - output_file: 输出文件路径
            - order_no: 订单号
            - store_code: 门店编号
            - warehouse_code: 仓库编码
            - item_count: 商品数量
            - unmatched_items: 未匹配SKU的商品列表
            - extracted_from: 输入来源（excel/image/pdf/text）
            - message: 消息
        """
        # ── v5.9.0 Phase 1：本次调用的 session_id + 反馈采集器初始化 ─────
        order_session_id = str(uuid.uuid4())
        _started_ms = int(time.time() * 1000)
        if _HAS_EVENT_BUS and get_feedback_collector() is None:
            try:
                init_feedback_collector(self.db_config)
            except Exception as _e:
                print(f"[WARN] init_feedback_collector failed: {_e}", flush=True)

        try:
            # 自动检测类型
            if order_type == "auto":
                order_type = object.__getattribute__(self, '_detect_input_type')(order_input)
            
            # 解析订单数据
            if order_type == "image":
                if ocr_result:
                    # 有OCR结果，直接处理
                    order_data = object.__getattribute__(self, '_handle_ocr_result')(ocr_result)
                    extracted_from = "image"
                elif order_input:
                    # 没有OCR结果，检查图片文件
                    parse_result = object.__getattribute__(self, '_parse_image')(order_input)
                    if parse_result and parse_result.get("__need_ocr__"):
                        # 返回需要OCR的标记
                        return {
                            "success": False,
                            "need_ocr": True,
                            "image_path": order_input,
                            "message": f"图片需要OCR识别，请稍候"
                        }
                    else:
                        order_data = parse_result
                        extracted_from = "image"
                else:
                    order_data = None
                    extracted_from = "image"
            elif order_type == "pdf":
                if ocr_result:
                    # 有OCR结果，直接处理
                    order_data = object.__getattribute__(self, '_handle_pdf_ocr_result')(ocr_result)
                    extracted_from = "pdf"
                elif order_input:
                    # 没有OCR结果，检查PDF文件
                    parse_result = object.__getattribute__(self, '_parse_pdf')(order_input)
                    if parse_result and parse_result.get("__need_pdf_ocr__"):
                        # 返回需要PDF OCR的标记
                        return {
                            "success": False,
                            "need_pdf_ocr": True,
                            "pdf_path": order_input,
                            "message": f"PDF需要OCR识别，请稍候"
                        }
                    else:
                        order_data = parse_result
                        extracted_from = "pdf"
                else:
                    order_data = None
                    extracted_from = "pdf"
            elif order_type == "text":
                # 【新增】优先使用 LLM 解析，失败则 fallback 到正则
                try:
                    llm_result = self.parse_with_llm(order_input, content_type="text")
                    if llm_result.get("success"):
                        order_data = object.__getattribute__(self, '_normalize_llm_result')(llm_result)
                        order_data["_parse_method"] = "llm"
                        order_data["_llm_confidence"] = llm_result.get("confidence", 0)
                    else:
                        raise ValueError(f"LLM解析失败: {llm_result.get('error', '未知错误')}")
                except Exception as llm_err:
                    # LLM 解析失败，fallback 到正则解析
                    order_data = object.__getattribute__(self, '_parse_text')(order_input)
                    if order_data:
                        order_data["_parse_method"] = "regex_fallback"
                        order_data["_parse_error"] = str(llm_err)
                extracted_from = "text"
            else:
                # excel - 【新增】优先使用 LLM 解析，失败则 fallback 到正则
                if not os.path.exists(order_input):
                    return {
                        **get_friendly_error("E001", f"文件不存在: {order_input}"),
                        "extracted_from": "excel"
                    }
                
                # Step 1: tools_parse() - 调用 tools/order_parser.parse()
                try:
                    parsed_result = self.tools_parse(order_input, order_type="excel")
                    if parsed_result.get("success"):
                        order_data = parsed_result
                        order_data["_parse_method"] = "tools_parse"
                    else:
                        # parse 失败，返回错误（不 fallback 到正则，parse() 内部会处理）
                        return {
                            **get_friendly_error("E101", parsed_result.get("error", "订单解析失败")),
                            "extracted_from": "excel"
                        }
                except Exception as parse_err:
                    return {
                        **get_friendly_error("E101", str(parse_err)),
                        "extracted_from": "excel"
                    }

                # Step 2: tools_transform() - 调用 field_transformer 规则库标准化
                try:
                    order_data = self.tools_transform(order_data)
                except Exception as transform_err:
                    # 规则库转换失败不影响主流程，记录 warning 继续
                    order_data.setdefault("_warnings", []).append(f"规则库转换异常: {str(transform_err)}")
                extracted_from = "excel"
            
            if not order_data:
                return get_friendly_error("E101", f"订单解析失败（来源：{order_type}）")

            # ========== 多门店处理 vs 单门店处理 ==========
            all_store_results = []  # [{store_info, sku_results, items}, ...]

            if order_data.get("_multi_store") and order_data.get("stores"):
                # 【多门店模式】遍历每个门店
                stores_dict = order_data["stores"]

                for store_key, store_data in stores_dict.items():
                    # 获取该门店的商品（带 _store_name 标记的）
                    store_items = [it for it in order_data["items"] 
                                   if it.get("_store_name") == store_key 
                                   or it.get("_store_name") == store_data.get("store_name", store_key)]
                    if not store_items:
                        store_items = store_data.get("items", [])
                        # 转换格式
                        store_items = [{
                            "seq": i + 1,
                            "product_code": str(it.get("product_code", "")).strip(),
                            "product_name": str(it.get("product_name", "")).strip(),
                            "spec": str(it.get("spec", "")).strip(),
                            "quantity": int(it.get("quantity", 0)),
                            "unit": str(it.get("unit", "件")).strip(),
                            "remark": str(it.get("remark", "")).strip(),
                        } for i, it in enumerate(store_items)]

                    store_name_for_match = store_data.get("store_name", store_key)

                    # 门店匹配：用户已确认门店时跳过匹配流程
                    if confirmed_store:
                        si = confirmed_store
                    else:
                        si = _call_match_store(
                            store_name=store_name_for_match,
                            customer_company=store_data.get("shipper_name", ""),
                            db_config=self.db_config,
                            phone=store_data.get("phone"),
                            address=store_data.get("store_address"),
                            contact_person=store_data.get("contact_person"),
                        )
                    if not si:
                        return {
                            **get_friendly_error("E201", f"门店「{store_name_for_match}」未找到匹配"),
                            "need_store_match": True,
                            "store_name_submitted": store_name_for_match
                        }
                    # 门店匹配后必须用户确认（含相似度信息）
                    if si.get("need_confirm") or si.get("candidates"):
                        top_c = si["candidates"][0] if si.get("candidates") else {}
                        top_sim = top_c.get("similarity", 0)
                        # v5.9.0 Phase 1：emit 门店需要确认事件
                        if _HAS_EVENT_BUS:
                            EventBus.emit("store_confirm_needed", {
                                "session_id": order_session_id,
                                "timestamp": time.time(),
                                "store_name_submitted": store_name_for_match,
                                "matched_store": {
                                    "store_code": si.get("store_code", top_c.get("store_code", "")),
                                    "store_name": si.get("store_name", top_c.get("store_name", "")),
                                    "owner_code": si.get("owner_code", top_c.get("owner_code", "")),
                                },
                                "candidates": si.get("candidates", []),
                                "top_similarity": top_sim,
                                "match_type": si.get("match_type", top_c.get("match_type", "")),
                                "match_layer": si.get("match_type", top_c.get("match_type", "")),
                                "need_customer_hint": False,
                            })
                        return {
                            "success": False,
                            "need_store_confirm": True,
                            "store_name_submitted": store_name_for_match,
                            "candidates": si.get("candidates", []),
                            "matched_store": {
                                "store_code": si.get("store_code", top_c.get("store_code", "")),
                                "store_name": si.get("store_name", top_c.get("store_name", "")),
                                "owner_code": si.get("owner_code", top_c.get("owner_code", "")),
                                "owner_name": si.get("owner_name", top_c.get("owner_name", "")),
                                "warehouse_name": si.get("warehouse_name", top_c.get("warehouse_name", "")),
                                "warehouse_code": si.get("warehouse_code", top_c.get("warehouse_code", "")),
                                "address": si.get("address", top_c.get("address", "")),
                                "contact_person": si.get("contact_person", top_c.get("contact_person", "")),
                                "phone": si.get("phone", top_c.get("phone", "")),
                                "similarity": top_sim,
                                "match_type": si.get("match_type", top_c.get("match_type", "")),
                                "match_method": si.get("match_method", ""),
                            },
                            "message": f"门店「{store_name_for_match}」→ {si.get('store_name', top_c.get('store_name', ''))}（相似度{top_sim:.0f}%），请确认"
                        }
                    if si.get("need_customer_hint"):
                        # 多候选时：顶候选相似度 >= 90% 则自动确认
                        top_c = si["candidates"][0]
                        top_sim = top_c.get("similarity", 0)
                        if top_sim >= 90.0 or top_sim >= 0.9:
                            c = top_c
                            si = {
                                "store_code": c["store_code"],
                                "store_name": c["store_name"],
                                "owner_code": c["owner_code"],
                                "owner_name": c.get("owner_name", ""),
                                "warehouse_name": c.get("warehouse_name", ""),
                                "warehouse_code": c.get("warehouse_code", ""),
                                "address": c.get("address", ""),
                                "contact_person": c.get("contact_person", ""),
                                "phone": c.get("phone", ""),
                                "match_type": "auto_confirm",
                                "match_method": f"自动确认（顶候选相似度{top_sim}）"
                            }
                        else:
                            return {
                                "success": False,
                                "need_store_confirm": True,
                                "store_name_submitted": store_name_for_match,
                                "candidates": si.get("candidates", []),
                                "message": f"门店「{store_name_for_match}」找到 {len(si.get('candidates', []))} 个候选门店（最高相似度{top_sim}）"
                            }
                    if si.get("need_customer_hint"):
                        return {
                            "success": False,
                            "need_customer_hint": True,
                            "store_name_submitted": store_name_for_match,
                            "possible_customers": si.get("possible_customers", []),
                            "message": f"门店「{store_name_for_match}」未找到匹配，但可能属于以下货主"
                        }

                    owner_code = si.get("owner_code", self.shipper_id)
                    sku_results, unmatched_items = object.__getattribute__(self, '_match_sku')(store_items, owner_code)

                    all_store_results.append({
                        "store_info": si,
                        "store_name": store_name_for_match,
                        "sku_results": sku_results,
                        "unmatched_items": unmatched_items,
                        "items": store_items,
                    })

            else:
                # 【单门店模式】
                # 从 stores 字典中提取单个门店信息（如果存在）
                stores_dict = order_data.get("stores", {})
                if stores_dict:
                    # 取第一个门店
                    store_key = list(stores_dict.keys())[0]
                    store_data = stores_dict[store_key]
                    store_name_val = store_data.get("store_name", store_key)
                    phone_val = store_data.get("phone")
                    address_val = store_data.get("address")
                    contact_val = store_data.get("contact_person")
                    shipper_name_val = store_data.get("shipper_name", "")
                else:
                    store_name_val = order_data.get("store_name", "")
                    phone_val = order_data.get("phone")
                    address_val = order_data.get("address")
                    contact_val = order_data.get("contact_person")
                    shipper_name_val = order_data.get("customer_company", "")

                if confirmed_store:
                    store_info = confirmed_store
                else:
                    store_info = _call_match_store(
                        store_name=store_name_val,
                        customer_company=shipper_name_val,
                        db_config=self.db_config,
                        phone=phone_val,
                        address=address_val,
                        contact_person=contact_val,
                    )

                    if not store_info:
                        return {
                            "success": False,
                            "need_store_match": True,
                            "store_name_submitted": order_data.get("store_name", ""),
                            "message": f"门店「{order_data.get('store_name', '未知')}」未找到匹配门店"
                        }
                    # ⚠️ 不再自动确认：所有门店匹配都必须用户确认
                    if store_info.get("need_confirm"):
                        # v5.9.0 Phase 1：emit 门店需要确认事件（单门店版）
                        if _HAS_EVENT_BUS:
                            top_c = (store_info.get("candidates") or [{}])[0]
                            top_sim = top_c.get("similarity", 0)
                            EventBus.emit("store_confirm_needed", {
                                "session_id": order_session_id,
                                "timestamp": time.time(),
                                "store_name_submitted": store_info.get("store_name_submitted", ""),
                                "matched_store": {
                                    "store_code": top_c.get("store_code", ""),
                                    "store_name": top_c.get("store_name", ""),
                                    "owner_code": top_c.get("owner_code", ""),
                                },
                                "candidates": store_info.get("candidates", []),
                                "top_similarity": top_sim,
                                "match_type": top_c.get("match_type", "unknown"),
                                "match_layer": top_c.get("match_type", "unknown"),
                                "need_customer_hint": False,
                            })
                        return {
                            "success": False,
                            "need_store_confirm": True,
                            "store_name_submitted": store_info.get("store_name_submitted", ""),
                            "candidates": store_info.get("candidates", []),
                            "message": f"门店「{store_info.get('store_name_submitted', '未知')}」找到 {len(store_info.get('candidates', []))} 个候选门店，请确认"
                        }
                    if store_info.get("need_customer_hint"):
                        return {
                            "success": False,
                            "need_customer_hint": True,
                            "store_name_submitted": store_info.get("store_name_submitted", ""),
                            "possible_customers": store_info.get("possible_customers", []),
                            "message": f"门店「{store_info.get('store_name_submitted', '未知')}」未找到匹配，但可能属于以下货主"
                        }

                owner_code = store_info.get("owner_code", self.shipper_id) if store_info else self.shipper_id
                sku_results, unmatched_items = object.__getattribute__(self, '_match_sku')(order_data["items"], owner_code)

                all_store_results.append({
                    "store_info": store_info,
                    "store_name": store_info.get("store_name", "") if store_info else "",
                    "sku_results": sku_results,
                    "unmatched_items": unmatched_items,
                    "items": order_data["items"],
                })

            # ========== 生成输出文件名 ==========
            if not output_file:
                raw_order_no = order_data.get("order_no", "")
                if raw_order_no and raw_order_no not in ("unknown", "", "nan", "往来单位名称", "单据日期"):
                    order_no_safe = raw_order_no.replace("/", "-").replace("\\", "-")
                else:
                    now_str = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                    order_no_safe = f"DH-O-{now_str}"
                output_file = os.path.join(self.output_dir, f"华鼎出库单_{order_no_safe}.xlsx")

            # ========== 生成合并模板（所有门店写入同一个sheet）==========
            object.__getattribute__(self, '_generate_multi_store_template')(order_data, all_store_results, output_file)

            # ========== 同步到 media/outbound（供下载链接访问）==========
            try:
                import shutil
                outbound_dir = os.path.expanduser("~/.openclaw/media/outbound")
                os.makedirs(outbound_dir, exist_ok=True)
                outbound_file = os.path.join(outbound_dir, os.path.basename(output_file))
                shutil.copy2(output_file, outbound_file)
            except Exception as copy_err:
                print(f"[WARN] 复制到 media/outbound 失败: {copy_err}")

            # ========== 统计汇总 ==========
            total_items = sum(len(r["sku_results"]) for r in all_store_results)
            total_unmatched = sum(len(r["unmatched_items"]) for r in all_store_results)
            all_unmatched = []
            for r in all_store_results:
                for it in r["unmatched_items"]:
                    it["_store"] = r["store_name"]
                    all_unmatched.append(it)

            # ========== 生成汇总映射对照表 ==========
            review_data = object.__getattribute__(self, '_generate_mapping_comparison_multi')(
                order_data=order_data,
                all_store_results=all_store_results,
            )

            has_issues = total_unmatched > 0 or review_data["summary"]["alert_count"] > 0
            store_names = ", ".join(r["store_name"] for r in all_store_results)

            # ========== 格式化用户友好的返回消息 ==========
            friendly_msg = object.__getattribute__(self, '_format_success_message')(
                store_names=store_names,
                item_count=total_items,
                matched_count=total_items - total_unmatched,
                unmatched_count=total_unmatched,
                file_name=os.path.basename(output_file),
                has_issues=has_issues
            )

            return {
                "success": True,
                "need_review": True,
                # 文件信息
                "output_file": output_file,
                "file_name": os.path.basename(output_file),
                # download_url: 基于 AWS 公网 IP 生成下载链接
                "download_url": _get_download_url(output_file),
                # 订单信息
                "order_no": order_data.get("order_no", ""),
                "store_names": store_names,
                "store_count": len(all_store_results),
                "item_count": total_items,
                # 匹配统计
                "matched_count": total_items - total_unmatched,
                "unmatched_count": total_unmatched,
                "unmatched_items": all_unmatched,
                # 详细数据
                "extracted_from": extracted_from,
                "review_data": review_data,
                "has_issues": has_issues,
                "all_store_results": all_store_results,
                # 用户消息
                "message": friendly_msg
            }

        except Exception as e:
            return {
                **wrap_error(e, "E999"),
                "_error_location": "execute()"
            }
    
    def _parse_order_excel(self, order_file: str) -> Optional[Dict]:
        """
        解析客户订单Excel - 按表头名读取，而非固定行列
        
        支持不同客户的不同Excel格式
        """
        try:
            import pandas as pd
            
            df_raw = pd.read_excel(order_file, header=None)
            warnings = []
            
            # ========== Step 1: 找到表头行，建立列映射 ==========
            # 表头可能在第0-6行（某些Excel表头比较靠前）
            col_mapping, header_warnings = self.normalize_excel_headers(df_raw, header_row_hints=[0, 1, 2, 3, 4, 5, 6])
            warnings.extend(header_warnings)
            # 建立反向映射：标准字段名 → 列索引
            field_to_col = {v: k for k, v in col_mapping.items()}
            
            # ========== Step 2: 提取元数据（门店、订单号等）============
            # 尝试从常见位置读取元数据（兼容旧格式）
            order_no = ""
            store_name = ""
            contact_person = ""
            phone = ""
            address = ""
            
            # 尝试从col_mapping读取
            if "order_no" in col_mapping:
                col_idx = col_mapping["order_no"]
                if len(df_raw) > 1:
                    order_no = str(df_raw.iloc[1, col_idx]) if pd.notna(df_raw.iloc[1, col_idx]) else ""
            
            if "store_name" in col_mapping:
                col_idx = col_mapping["store_name"]
                if len(df_raw) > 2:
                    store_name = str(df_raw.iloc[2, col_idx]) if pd.notna(df_raw.iloc[2, col_idx]) else ""
            
            if "contact_person" in col_mapping:
                col_idx = col_mapping["contact_person"]
                if len(df_raw) > 2:
                    contact_person = str(df_raw.iloc[2, col_idx]) if pd.notna(df_raw.iloc[2, col_idx]) else ""
            
            if "phone" in col_mapping:
                col_idx = col_mapping["phone"]
                if len(df_raw) > 3:
                    phone = str(df_raw.iloc[3, col_idx]) if pd.notna(df_raw.iloc[3, col_idx]) else ""
            
            # Fallback: 如果col_mapping中没有store_name等字段，直接从固定位置读取
            if not store_name and len(df_raw) > 2 and len(df_raw.columns) > 1:
                # 门店名通常在 row=2, col=1
                val = df_raw.iloc[2, 1]
                if pd.notna(val):
                    store_name = str(val).strip()
            
            if not order_no and len(df_raw) > 1 and len(df_raw.columns) > 1:
                # 订单号通常在 row=1, col=1
                val = df_raw.iloc[1, 1]
                if pd.notna(val):
                    order_no = str(val).strip()
            
            if not contact_person and len(df_raw) > 2:
                # 联系人通常在 row=2, col=6
                val = df_raw.iloc[2, 6]
                if pd.notna(val):
                    contact_person = str(val).strip()
            
            if not phone and len(df_raw) > 3:
                # 电话通常在 row=3, col=6
                val = df_raw.iloc[3, 6]
                if pd.notna(val):
                    phone = str(val).strip()
            
            # 清理门店名称（去掉前缀）
            for prefix in ['河北-', '天津-', '沧州-']:
                if store_name.startswith(prefix):
                    store_name = store_name.replace(prefix, "")
                    break
            
            # ========== Step 3: 提取商品明细 ==========
            items = []
            data_start_row = 6  # 默认数据开始行
            
            # 找到表头行后，确定数据开始行
            if col_mapping:
                # 找到表头行的位置（同 header_row_hints 范围）
                for row_idx in [0, 1, 2, 3, 4, 5, 6]:
                    if row_idx >= len(df_raw):
                        continue
                    row = df_raw.iloc[row_idx]
                    has_header = any(
                        pd.notna(cell) and self.find_standard_field(str(cell))
                        for cell in row
                    )
                    if has_header:
                        data_start_row = row_idx + 1
                        break
            
            # 用 col_mapping（字段名→列索引）读取商品明细，兜底硬编码
            for i in range(data_start_row, len(df_raw)):
                seq = df_raw.iloc[i, 0] if pd.notna(df_raw.iloc[i, 0]) else None
                
                # 如果是合计行，跳过
                if seq is not None and (str(seq) == "合计：" or str(seq).startswith("合计")):
                    continue
                
                # 用 field_to_col 获取列索引，找不到则用硬编码兜底
                col_product_name = field_to_col.get("product_name", 4)
                col_spec         = field_to_col.get("spec", 5)
                col_quantity     = field_to_col.get("quantity", 7)
                col_unit         = field_to_col.get("unit", 6)
                col_remark       = field_to_col.get("remark", 8)
                col_product_code = field_to_col.get("product_code", 1)
                
                # 读取各字段
                product_name = str(df_raw.iloc[i, col_product_name]).strip() \
                    if pd.notna(df_raw.iloc[i, col_product_name]) else ""
                
                # 跳过表头行
                if product_name in ["商品名称", "商品编码", "商品规格", "计量单位", "备注", ""]:
                    continue
                if not product_name or product_name == "nan":
                    continue
                
                product_code = str(df_raw.iloc[i, col_product_code]).strip() \
                    if pd.notna(df_raw.iloc[i, col_product_code]) else ""
                spec = str(df_raw.iloc[i, col_spec]).strip() \
                    if pd.notna(df_raw.iloc[i, col_spec]) else ""
                
                qty_raw = df_raw.iloc[i, col_quantity]
                quantity = int(qty_raw) if pd.notna(qty_raw) and isinstance(qty_raw, (int, float)) else 0
                
                unit = str(df_raw.iloc[i, col_unit]).strip() \
                    if pd.notna(df_raw.iloc[i, col_unit]) else "件"
                remark = str(df_raw.iloc[i, col_remark]).strip() \
                    if pd.notna(df_raw.iloc[i, col_remark]) else ""
                
                items.append({
                    "seq": int(seq) if isinstance(seq, (int, float)) else len(items) + 1,
                    "product_code": product_code,
                    "product_name": product_name,
                    "spec": spec,
                    "quantity": quantity,
                    "unit": unit,
                    "remark": remark
                })
            
            if not store_name and not items:
                warnings.append("未能从Excel中提取到门店名和商品明细")
            
            return {
                "order_no": order_no,
                "store_name": store_name,
                "contact_person": contact_person,
                "phone": phone,
                "address": address,
                "items": items,
                "_norm_warnings": warnings
            }
            
        except Exception as e:
            print(f"订单解析错误: {e}")
            return None
    
    def _match_store(self, store_name: str, customer_company: str = None) -> Optional[Dict]:
        """
        门店匹配 - 三层匹配 + 货主提示策略
        
        第0层：客户公司匹配（优先）- 用customer_company直接查customer表拿到owner_code
        第1层：精确匹配（store_name = 输入名称）
        第2层：品牌+仓库匹配（精确匹配失败后）
        第3层：模糊匹配（相似度计算），相似度<75%视为匹配不到
        
        Args:
            store_name: 门店名称
            customer_company: 客户公司名称（OCR识别到的货主公司名，如"潍坊膳仪食品有限公司"）
            
        Returns:
            精确匹配时：返回门店信息 dict
            模糊匹配时：返回 {"need_confirm": True, "candidates": [...], "store_name_submitted": ...}
            匹配不到但有相似货主：返回 {"need_customer_hint": True, "possible_customers": [...]}
            匹配不到时：返回 None
        """
        try:
            import psycopg2
            from difflib import SequenceMatcher
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            # ========== 第0层：客户公司匹配（优先）==========
            if customer_company:
                # 直接用客户公司名查customer表，获取owner_code
                cur.execute("""
                    SELECT customer_id, customer_name, warehouse_name
                    FROM customer
                    WHERE status = 'ACTIVE' AND customer_name LIKE %s
                    LIMIT 5
                """, (f"%{customer_company}%",))
                company_rows = cur.fetchall()
                
                if company_rows:
                    matched_owner = company_rows[0][0]
                    company_name = company_rows[0][1]
                    
                    # 用owner_code查store_list，找第一个门店
                    cur.execute("""
                        SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                        FROM store_list
                        WHERE owner_code = %s
                        LIMIT 1
                    """, (matched_owner,))
                    row = cur.fetchone()
                    
                    if row:
                        warehouse_name = row[2] or ""
                        conn.close()
                        return {
                            "store_code": row[0] or "",
                            "store_name": row[1] or "",
                            "warehouse_name": warehouse_name,
                            "warehouse_code": object.__getattribute__(self, '_get_warehouse_code')(warehouse_name),
                            "address": row[3] or "",
                            "contact_person": row[4] or "",
                            "phone": row[5] or "",
                            "owner_code": row[6] or "",
                            "match_type": "customer_company",
                            "match_method": f"客户公司匹配（'{customer_company}'→'{company_name}'）"
                        }
                    
                    # 【新增】customer_company匹配到货主但store_list没有门店时
                    # 用品牌名查customer表，找品牌对应的owner_code，再走后续匹配
                    if store_name:
                        # 提取品牌部分（如"大铁牛郑州仓" → "大铁牛"）
                        brand = store_name
                        for suffix in ['仓', '仓店', '仓储']:
                            if suffix in store_name:
                                idx = store_name.rfind(suffix)
                                if idx > 0:
                                    brand = store_name[:idx]
                                    break
                        
                        # 用品牌名去customer表查找货主
                        cur.execute("""
                            SELECT customer_id, customer_name
                            FROM customer
                            WHERE status = 'ACTIVE' AND customer_name LIKE %s
                            LIMIT 5
                        """, (f"%{brand}%",))
                        brand_rows = list(cur.fetchall())
                        
                        if brand_rows:
                            # 用品牌匹配到的owner覆盖matched_owner
                            matched_owner = brand_rows[0][0]
                            company_name = brand_rows[0][1]
                            
                            # 用新的owner_code查store_list
                            cur.execute("""
                                SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                                FROM store_list
                                WHERE owner_code = %s
                                LIMIT 1
                            """, (matched_owner,))
                            row = cur.fetchone()
                            
                            if row:
                                warehouse_name = row[2] or ""
                                conn.close()
                                return {
                                    "store_code": row[0] or "",
                                    "store_name": row[1] or "",
                                    "warehouse_name": warehouse_name,
                                    "warehouse_code": object.__getattribute__(self, '_get_warehouse_code')(warehouse_name),
                                    "address": row[3] or "",
                                    "contact_person": row[4] or "",
                                    "phone": row[5] or "",
                                    "owner_code": row[6] or "",
                                    "match_type": "brand_mapping",
                                    "match_method": f"品牌映射匹配（'{brand}'→'{company_name}'）"
                                }
            
            def calc_similarity(s1, s2):
                """计算两个字符串的相似度 (0-1)
                
                策略：
                1. 先做整体相似度（SequenceMatcher）
                2. 如果短串是长串的前缀/包含，额外加分
                3. 清理品牌前缀后再比较
                4. 核心关键词匹配（地名/店名等）加分
                """
                if not s1 or not s2:
                    return 0.0
                
                # 品牌前缀列表（按长度降序排列，避免误匹配）
                prefixes = ['制茶青年-', '制茶青年', '天津仓-', '天津仓']
                
                # 清理品牌前缀
                s1_cleaned = s1
                s2_cleaned = s2
                for prefix in prefixes:
                    if s1_cleaned.startswith(prefix):
                        s1_cleaned = s1_cleaned[len(prefix):]
                    if s2_cleaned.startswith(prefix):
                        s2_cleaned = s2_cleaned[len(prefix):]
                
                # 1. 基础相似度
                base_sim = SequenceMatcher(None, s1_cleaned, s2_cleaned).ratio()
                
                # 2. 包含匹配加分（短串是长串的前缀/包含）
                # 确保s1是较短的串
                shorter = s1_cleaned if len(s1_cleaned) <= len(s2_cleaned) else s2_cleaned
                longer = s2_cleaned if len(s1_cleaned) <= len(s2_cleaned) else s1_cleaned
                
                contain_bonus = 0.0
                if shorter and shorter in longer:
                    # 包含匹配相似度 = 短串长度/长串长度 + 0.3
                    contain_bonus = len(shorter) / len(longer) + 0.3
                
                # 3. 核心关键词匹配（地名关键词）
                # 提取关键位置词：省/市/区/县/镇/乡/村/店等
                location_pattern = r'[\u4e00-\u9fa5]{2,6}(?:省|市|区|县|镇|乡|村|街|路|店|仓|口)'
                import re
                loc1 = set(re.findall(location_pattern, s1_cleaned))
                loc2 = set(re.findall(location_pattern, s2_cleaned))
                
                keyword_bonus = 0.0
                if loc1 and loc2:
                    # 有共同关键词则加分
                    common = loc1 & loc2
                    if common:
                        # 每个共同关键词加0.2，最高0.6
                        keyword_bonus = min(len(common) * 0.2, 0.6)
                
                return max(base_sim, contain_bonus, base_sim + keyword_bonus)
            
            def try_exact_match(name):
                """精确匹配：store_name = 输入名称"""
                cur.execute("""
                    SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                    FROM store_list
                    WHERE store_name = %s
                    LIMIT 1
                """, (name,))
                return cur.fetchone()
            
            def try_brand_warehouse_match(name):
                """
                【新增】品牌+仓库名匹配策略
                
                逻辑：
                1. 从门店名提取"完整仓库名称"（如"大铁牛郑州仓" → 仓库="郑州仓"）
                2. 提取"品牌部分"（如"大铁牛郑州仓" → 品牌="大铁牛"）
                3. 用仓库名称精确查store_list的warehouse字段（warehouse = '郑州仓'）
                4. 用品牌名去customer表查找货主ID
                5. 用仓库名 + 货主 去查store_list，找到真实门店名
                
                Args:
                    name: 门店名称（如"大铁牛郑州仓"）
                Returns:
                    匹配到的门店信息，或 None
                """
                import re
                
                # 仓库关键词后缀
                warehouse_suffixes = ['仓', '仓店', '仓储']
                
                # 提取完整仓库名称（如"郑州仓"而不是"仓"）
                warehouse_part = None
                brand_part = name
                
                for suffix in warehouse_suffixes:
                    if suffix in name:
                        idx = name.rfind(suffix)
                        if idx > 0:
                            # 往前找完整的仓库名称（2-4个字符+后缀）
                            # 如"大铁牛郑州仓" → "郑州仓"
                            start = max(0, idx - 4)
                            potential = name[start:idx + len(suffix)]
                            
                            # 确保仓库名以仓/店结尾
                            if potential.endswith('仓') or potential.endswith('店'):
                                warehouse_part = potential
                                brand_part = name[:start] if start > 0 else name[:idx]
                                break
                
                if not warehouse_part:
                    return None
                
                # 【修复】精确匹配warehouse字段 = '郑州仓'（不是 LIKE '%仓%'）
                cur.execute("""
                    SELECT DISTINCT owner_code, warehouse
                    FROM store_list
                    WHERE warehouse = %s AND owner_code IS NOT NULL AND owner_code != ''
                """, (warehouse_part,))
                warehouse_owners = list(cur.fetchall())
                
                if not warehouse_owners:
                    return None
                
                # 用品牌名去customer表查找货主
                cur.execute("""
                    SELECT customer_id, customer_name
                    FROM customer
                    WHERE status = 'ACTIVE' AND customer_name LIKE %s
                """, (f"%{brand_part}%",))
                brand_customers = list(cur.fetchall())
                
                if not brand_customers:
                    # 如果customer表没找到，用owner_code去customer表反向查
                    owner_ids = [r[0] for r in warehouse_owners]
                    cur.execute("""
                        SELECT customer_id, customer_name
                        FROM customer
                        WHERE status = 'ACTIVE' AND customer_id = ANY(%s)
                    """, (owner_ids,))
                    brand_customers = list(cur.fetchall())
                
                if not brand_customers:
                    return None
                
                # 找到匹配的货主ID
                matched_owner = brand_customers[0][0]
                
                # 【修复】精确匹配 warehouse = '郑州仓' + 货主ID 查store_list
                cur.execute("""
                    SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                    FROM store_list
                    WHERE warehouse = %s AND owner_code = %s
                    LIMIT 5
                """, (warehouse_part, matched_owner))
                rows = cur.fetchall()
                
                if not rows:
                    return None
                
                # 返回第一个匹配结果
                row = rows[0]
                return row
                
                if not rows:
                    return None
                
                # 返回第一个匹配结果（也可以按相似度排序后选最佳）
                row = rows[0]
                return row
            
            def try_fuzzy_match(name, top_n=5):
                """模糊匹配：返回多个候选，按相似度排序
                
                策略：
                1. 先用完整名称LIKE查询
                2. 如果没有结果，拆分为关键词逐个查询后合并去重
                """
                all_rows = []
                
                # 第1步：完整名称LIKE
                cur.execute("""
                    SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                    FROM store_list
                    WHERE store_name LIKE %s
                    LIMIT 50
                """, (f"%{name}%",))
                all_rows = list(cur.fetchall())
                
                # 第2步：如果没有结果，提取关键词分别查询
                if not all_rows:
                    # 提取2-4个字的中文关键词
                    import re
                    keywords = re.findall(r'[\u4e00-\u9fa5]{2,4}', name)
                    seen_codes = set()
                    for kw in keywords:
                        if len(kw) >= 2:
                            cur.execute("""
                                SELECT store_code, store_name, warehouse, address, contact_person, phone, owner_code
                                FROM store_list
                                WHERE store_name LIKE %s
                                LIMIT 20
                            """, (f"%{kw}%",))
                            for row in cur.fetchall():
                                if row[0] not in seen_codes:
                                    all_rows.append(row)
                                    seen_codes.add(row[0])
                
                # 计算相似度并排序
                scored = []
                for row in all_rows:
                    db_name = row[1] or ""
                    sim = calc_similarity(name, db_name)
                    scored.append((sim, row))
                
                scored.sort(key=lambda x: x[0], reverse=True)
                return scored[:top_n]
            
            def search_similar_customers(name):
                """搜索customer表查找相似货主名"""
                cur.execute("""
                    SELECT customer_id, customer_name, warehouse_name, status
                    FROM customer
                    WHERE status = 'ACTIVE'
                    LIMIT 20
                """)
                rows = cur.fetchall()
                
                # 计算相似度
                scored = []
                for row in rows:
                    cust_name = row[1] or ""
                    # 相似度：门店名 vs 货主名
                    sim = calc_similarity(name, cust_name)
                    scored.append((sim, row))
                
                scored.sort(key=lambda x: x[0], reverse=True)
                # 只返回相似度>=75%的
                return [(sim, row) for sim, row in scored if sim >= 0.70]
            
            # ========== 第1层：精确匹配 ==========
            row = try_exact_match(store_name)
            
            if row:
                # 精确匹配成功，直接返回
                warehouse_name = row[2] or ""
                conn.close()
                return {
                    "store_code": row[0] or "",
                    "store_name": row[1] or "",
                    "warehouse_name": warehouse_name,
                    "warehouse_code": object.__getattribute__(self, '_get_warehouse_code')(warehouse_name),
                    "address": row[3] or "",
                    "contact_person": row[4] or "",
                    "phone": row[5] or "",
                    "owner_code": row[6] or "",
                    "match_type": "exact",
                    "match_method": "精确匹配"
                }
            
            # ========== 第2层：品牌+仓库名匹配（精确匹配失败后）==========
            row = try_brand_warehouse_match(store_name)
            
            if row:
                warehouse_name = row[2] or ""
                conn.close()
                return {
                    "store_code": row[0] or "",
                    "store_name": row[1] or "",
                    "warehouse_name": warehouse_name,
                    "warehouse_code": object.__getattribute__(self, '_get_warehouse_code')(warehouse_name),
                    "address": row[3] or "",
                    "contact_person": row[4] or "",
                    "phone": row[5] or "",
                    "owner_code": row[6] or "",
                    "match_type": "brand_warehouse",
                    "match_method": f"品牌+仓库匹配（门店'{store_name}'→仓库'{warehouse_name}'）"
                }
            
            # ========== 第3层：模糊匹配（品牌+仓库匹配失败后）==========
            candidates = try_fuzzy_match(store_name)
            
            # 过滤：相似度 < 75% 的结果视为匹配不到
            high_confidence = [(sim, row) for sim, row in candidates if sim >= 0.70]
            
            if high_confidence:
                # 返回相似度 >= 75% 的候选列表（需要用户确认）
                conn.close()
                
                candidate_list = []
                for sim, row in high_confidence:
                    warehouse_name = row[2] or ""
                    candidate_list.append({
                        "store_code": row[0] or "",
                        "store_name": row[1] or "",
                        "warehouse_name": warehouse_name,
                        "warehouse_code": object.__getattribute__(self, '_get_warehouse_code')(warehouse_name),
                        "address": row[3] or "",
                        "contact_person": row[4] or "",
                        "phone": row[5] or "",
                        "owner_code": row[6] or "",
                        "similarity": round(sim * 100, 1)
                    })
                
                return {
                    "need_confirm": True,
                    "candidates": candidate_list,
                    "store_name_submitted": store_name,
                    "match_type": "fuzzy",
                    "match_method": f"模糊匹配({len(candidate_list)}个候选)"
                }
            
            # ========== 第3层：门店匹配失败，搜索相似货主 ==========
            similar_customers = search_similar_customers(store_name)
            
            conn.close()
            
            if similar_customers:
                customer_list = []
                for sim, row in similar_customers:
                    customer_list.append({
                        "customer_id": row[0] or "",
                        "customer_name": row[1] or "",
                        "warehouse_name": row[2] or "",
                        "similarity": round(sim * 100, 1)
                    })
                
                return {
                    "need_customer_hint": True,
                    "possible_customers": customer_list,
                    "store_name_submitted": store_name,
                    "match_type": "customer_hint",
                    "match_method": f"货主提示({len(customer_list)}个候选)"
                }
            
            # 真的匹配不到
            return None
            
        except Exception as e:
            print(f"门店匹配错误: {e}")
            return None

    def _clean_product_name(self, name: str) -> str:
        """
        清理商品名称，去除特殊字符用于匹配
        
        处理规则：
        - 去除首尾特殊字符（-、/、.等）
        - 去除括号及其内容
        - 去除空格
        """
        import re
        # 去除首尾的特殊字符
        name = name.strip().rstrip('-').rstrip('/').rstrip('.')
        # 去除括号及其内容
        name = re.sub(r'[（(].*[)）]', '', name)
        # 去除空格
        name = name.strip()
        # 去除名称中间的"-"、"（）"等特殊符号（只保留中文、字母、数字、下划线、/）
        name = re.sub(r'[^\u4e00-\u9fff\w/]', '', name)
        return name
    
    def _match_sku(self, items: list, owner_code: str) -> tuple:
        """
        SKU匹配 - 调用 sku_mapper.map_sku_batch 批量模式（2026-06-03 优化）

        优化前：每个商品一次 DB 查询，481商品约2400次查询 → 几分钟
        优化后：每个货主2次 DB 查询（SKU缓存+别名表），剩余全内存匹配 → 几十毫秒

        Args:
            items: 订单商品列表
            owner_code: 货主ID

        Returns:
            (results, unmatched_items)
        """
        from skills.skill_order_to_huading_template.tools._sku_mapper import map_sku_batch

        # 批量处理，一次 DB 查询获取所有 SKU，后续全内存匹配
        results, unmatched_items = map_sku_batch(
            owner_code, items, self.db_config
        )

        # 转换格式，与旧接口兼容
        formatted_results = []
        for r in results:
            # 找到原始 item 的 seq
            seq = next((item.get("seq", 0) for item in items
                        if item.get("product_name", "") == r.get("product_name", "")
                        or item.get("product_name", "") == r.get("sku_name", "")
                        ), 0)
            formatted_results.append({
                "seq": seq,
                "product_name": r["product_name"],
                "spec": r.get("spec", ""),
                "quantity": r.get("quantity", 0),
                "remark": r.get("remark", ""),
                "sku_code": r["sku_code"],
                "sku_name": r["sku_name"],
                "unit": r["unit"],
                "unit_type": r["unit_type"],
                "product_spec": r.get("product_spec", ""),
                "match_method": r.get("match_method", ""),
            })

        return formatted_results, unmatched_items

    def _get_sku_unit_info(self, sku_code: str, owner_code: str) -> Dict[str, str]:
        """
        获取SKU的单位信息
        
        Returns:
            dict: {
                "unit": "件/箱/袋",  # 华鼎单位（具体单位名称）
                "unit_type": "大单位/中单位/小单位"
            }
        """
        try:
            import psycopg2
            import json
            
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()
            
            # Step 1: 用 sku_code 找到该SKU对应的商品名称
            cur.execute("""
                SELECT m.system_sku_name
                FROM shipper_sku_mapping m
                WHERE m.shipper_id = %s AND m.system_sku_code = %s AND m.status = 'ACTIVE'
                LIMIT 1
            """, (owner_code, sku_code))
            row = cur.fetchone()
            if not row:
                conn.close()
                return {"unit": "件", "unit_type": "大单位"}
            
            product_name = row[0]
            
            # Step 2: 找到该商品名称对应的所有系统SKU及其单位配置
            cur.execute("""
                SELECT m.system_sku_code, m.unit_conversion_rule
                FROM shipper_sku_mapping m
                WHERE m.shipper_id = %s 
                  AND m.system_sku_name = %s 
                  AND m.status = 'ACTIVE'
            """, (owner_code, product_name))
            all_rows = cur.fetchall()
            conn.close()
            
            if not all_rows:
                return {"unit": "件", "unit_type": "大单位"}
            
            # 提取所有SKU的单位配置: {sku_code: (ratio, unit), ...}
            sku_units = {}
            for r in all_rows:
                s_code = r[0]
                rule = r[1]
                if not rule:
                    continue
                
                if isinstance(rule, str):
                    rule = json.loads(rule)
                
                if isinstance(rule, dict):
                    ratio = rule.get('ratio')
                    unit = rule.get('unit')
                    if ratio and unit:
                        sku_units[s_code] = (float(ratio), unit)
            
            if not sku_units:
                return {"unit": "件", "unit_type": "大单位"}
            
            # 获取当前sku的单位信息
            current_unit_info = sku_units.get(sku_code)
            if not current_unit_info:
                return {"unit": "件", "unit_type": "大单位"}
            
            current_ratio, current_unit = current_unit_info
            
            # 只有一个SKU配置时，默认返回"大单位"
            if len(sku_units) == 1:
                return {"unit": current_unit, "unit_type": "大单位"}
            
            # 多个SKU时，按ratio排序判断单位类型
            sorted_skus = sorted(sku_units.items(), key=lambda x: x[1][0])
            
            if len(sorted_skus) == 2:
                if current_ratio == sorted_skus[1][1][0]:  # ratio最大
                    return {"unit": current_unit, "unit_type": "大单位"}
                else:
                    return {"unit": current_unit, "unit_type": "小单位"}
            
            elif len(sorted_skus) >= 3:
                min_ratio = sorted_skus[0][1][0]
                max_ratio = sorted_skus[-1][1][0]
                
                if current_ratio == max_ratio:
                    return {"unit": current_unit, "unit_type": "大单位"}
                elif current_ratio == min_ratio:
                    return {"unit": current_unit, "unit_type": "小单位"}
                else:
                    return {"unit": current_unit, "unit_type": "中单位"}
            
            return {"unit": current_unit, "unit_type": "大单位"}
                
        except Exception as e:
            print(f"获取单位信息失败: {e}")
            return {"unit": "件", "unit_type": "大单位"}
    
    def _get_unit_type(self, sku_code: str, owner_code: str) -> str:
        """
        获取SKU的单位类型（大/中/小单位）
        
        判断逻辑：
        1. 用 sku_code 找到该SKU对应的商品名称
        2. 找到该商品名称对应的所有系统SKU及其单位配置
        3. 按 ratio 判断当前 sku_code 是大/中/小单位：
           - ratio最大的 → 大单位
           - ratio中等的 → 中单位
           - ratio最小的 → 小单位
        4. 如果只有1个单位配置，默认返回"大单位"
        
        注意：这里的"商品"是指系统SKU，同一个商品可能对应多个系统SKU
        （如原味晶球有袋装和箱装两种规格），需要按ratio判断单位类型。
        """
        return object.__getattribute__(self, '_get_sku_unit_info')(sku_code, owner_code)["unit_type"]

    
    def _generate_mapping_comparison(self, order_data: Dict, store_info: Dict,
                                      sku_results: list, unmatched_items: list) -> Dict[str, Any]:
        """
        生成映射对照表（人工审核用）
        
        字段对照：
        - 客户: 商品名称、规格、单位、数量
        - 华鼎: SKU编码、SKU名称、数量、单位、单位类型
        
        Args:
            order_data: 订单数据（包含items）
            store_info: 门店信息
            sku_results: SKU匹配结果列表
            unmatched_items: 未匹配商品列表
            
        Returns:
            {
                "success": True,
                "comparison_table": [...],  # 映射对照表数据
                "alerts": [...],  # 告警列表
                "summary": {...}  # 汇总信息
            }
        """
        comparison_table = []
        alerts = []
        
        # 获取订单商品列表
        order_items = order_data.get("items", [])
        
        # 构建order_items索引（按seq查找）
        order_dict = {item.get("seq", i+1): item for i, item in enumerate(order_items)}
        
        # 已匹配的商品
        for result in sku_results:
            seq = result["seq"]
            order_item = order_dict.get(seq, {})
            
            row = {
                "seq": seq,
                # 客户订单
                "customer_product_name": order_item.get("product_name", result.get("product_name", "")),  # 客户商品名称
                "customer_spec": order_item.get("spec", result.get("spec", "")),  # 客户规格
                "customer_unit": order_item.get("unit", "件"),  # 客户单位
                "customer_quantity": order_item.get("quantity", result.get("quantity", 1)),  # 客户数量
                # 华鼎订单
                "sku_code": result["sku_code"],  # 华鼎SKU编码
                "sku_name": result["sku_name"],  # SKU名称（华鼎商品名称）
                "huading_quantity": result.get("quantity", 1),  # 华鼎数量（通常与客户数量一致）
                "huading_unit": result.get("unit", "件"),  # 华鼎单位
                "unit_type": result.get("unit_type", "大单位"),  # 单位类型（大/中/小单位）
                # 状态
                "matched": True,
                "confidence": result.get("confidence", 1.0),
                "match_method": result.get("match_method", "")
            }
            comparison_table.append(row)
            
            # 检查告警条件
            # 1. 置信度<80%
            if row["confidence"] < 0.8:
                alerts.append({
                    "type": "low_confidence",
                    "row": seq,
                    "severity": "⚠️",
                    "message": f"第{seq}行SKU匹配置信度{row['confidence']:.0%}，低于80%",
                    "details": {
                        "customer_product_name": row["customer_product_name"],
                        "sku_code": row["sku_code"],
                        "confidence": row["confidence"]
                    }
                })
        
        # 未匹配的商品
        for item in unmatched_items:
            seq = item.get("seq", 0)
            row = {
                "seq": seq,
                # 客户订单
                "customer_product_name": item.get("product_name", ""),
                "customer_spec": item.get("spec", ""),
                "customer_unit": item.get("unit", "件"),
                "customer_quantity": item.get("quantity", 1),
                # 华鼎订单
                "sku_code": "",
                "sku_name": "",
                "huading_quantity": "",
                "huading_unit": "",
                "unit_type": "",
                # 状态
                "matched": False,
                "confidence": 0,
                "match_method": "未匹配"
            }
            comparison_table.append(row)
            
            alerts.append({
                "type": "unmatched",
                "row": seq,
                "severity": "🔴",
                "message": f"第{seq}行商品未匹配到系统SKU",
                "details": {
                    "customer_product_name": row["customer_product_name"],
                    "customer_code": item.get("product_code", "")
                }
            })
        
        # 按seq排序
        comparison_table.sort(key=lambda x: x["seq"])
        
        # 检查必填字段为空的情况（已匹配的行）
        for row in comparison_table:
            if not row["matched"]:
                continue
            
            # 检查华鼎单位是否为空
            if not row["huading_unit"]:
                alerts.append({
                    "type": "empty_field",
                    "row": row["seq"],
                    "severity": "🔴",
                    "message": f"第{row['seq']}行华鼎单位为空",
                    "details": {
                        "customer_product_name": row["customer_product_name"],
                        "field": "huading_unit"
                    }
                })
        
        # 汇总
        total_items = len(comparison_table)
        matched_count = len([r for r in comparison_table if r["matched"]])
        unmatched_count = total_items - matched_count
        alert_count = len(alerts)
        
        return {
            "success": True,
            "comparison_table": comparison_table,
            "alerts": alerts,
            "summary": {
                "total_items": total_items,
                "matched_count": matched_count,
                "unmatched_count": unmatched_count,
                "alert_count": alert_count,
                "has_critical_alerts": any(a["type"] in ["unmatched", "empty_field"] for a in alerts)
            }
        }
    
    def _format_comparison_table_text(self, mapping_result: Dict[str, Any], order_info: Dict = None) -> str:
        """
        格式化映射对照表为文本（用于飞书展示）
        
        字段对照：
        - 客户: 商品名称、商品规格、单位、数量
        - 华鼎: SKU编码、SKU名称、数量、单位、单位类型
        
        格式：
        | # | 客户商品名称 | 商品规格 | 单位 | 数量 | → | 华鼎SKU编码 | SKU名称 | 数量 | 单位 | 单位类型 | 状态 |
        """
        table = mapping_result["comparison_table"]
        alerts = mapping_result["alerts"]
        summary = mapping_result["summary"]
        
        lines = []
        
        # 添加订单信息
        if order_info:
            lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            lines.append("订单信息")
            lines.append(f"  订单号: {order_info.get('order_no', '-')}")
            lines.append(f"  门店: {order_info.get('store_name', '-')}")
            lines.append(f"  货主: {order_info.get('owner_name', '-')}")
            lines.append(f"  仓库: {order_info.get('warehouse_name', '-')}")
            lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            lines.append("")
        
        # 添加告警汇总
        if alerts:
            lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            lines.append(f"告警 ({len(alerts)} 项)")
            for a in alerts:
                lines.append(f"  {a['severity']} {a['message']}")
            lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            lines.append("")
        
        # 添加汇总
        status = "✅" if not summary["has_critical_alerts"] else "⚠️ 需要检查"
        lines.append(f"商品映射对照表（{summary['total_items']}条）{status}")
        lines.append(f"  已匹配: {summary['matched_count']} | 未匹配: {summary['unmatched_count']} | 告警: {summary['alert_count']}")
        lines.append("")
        
        # 表头 - 按新字段格式
        # 客户: 商品名称、商品规格、单位、数量
        # 华鼎: SKU编码、SKU名称、数量、单位、单位类型
        header = "| # | 客户商品名称 | 规格 | 单位 | 数量 | → | 华鼎SKU编码 | SKU名称 | 数量 | 单位 | 单位类型 | 状态|"
        separator = "|---|-------------|------|------|------|---|-------------|---------|------|------|----------|------|"
        lines.append(header)
        lines.append(separator)
        
        # 找出有告警的行号
        alert_rows = {}
        for a in alerts:
            alert_rows[a["row"]] = a
        
        # 数据行
        for row in table:
            seq = row["seq"]
            
            # 状态判断
            if not row["matched"]:
                status_icon = "🔴"
            elif row["confidence"] < 0.8:
                status_icon = "⚠️"
            else:
                status_icon = "✅"
            
            # 华鼎SKU为空时显示-
            sku_code = row["sku_code"] or "-"
            sku_name = row["sku_name"] or "-"
            huading_quantity = row["huading_quantity"] if row["huading_quantity"] != "" else "-"
            huading_unit = row["huading_unit"] or "-"
            unit_type = row["unit_type"] or "-"
            
            line = f"| {seq} | {row['customer_product_name']} | {row['customer_spec']} | {row['customer_unit']} | {row['customer_quantity']} | → | {sku_code} | {sku_name} | {huading_quantity} | {huading_unit} | {unit_type} | {status_icon}|"
            lines.append(line)
        
        lines.append("")
        lines.append("请输入指令：")
        lines.append("  - 确认 / 没问题 / 通过 → 生成模板")
        lines.append("  - 修改第X行 → 修改指定行")
        lines.append("  - 取消 → 放弃订单")
        
        return "\n".join(lines)
    
    def _parse_user_feedback(self, user_message: str, mapping_result: Dict[str, Any], 
                            current_order_data: Dict) -> Dict[str, Any]:
        """
        使用LLM解析用户反馈，返回结构化的修改指令
        
        Args:
            user_message: 用户的反馈消息
            mapping_result: 当前映射结果
            current_order_data: 当前订单数据
            
        Returns:
            {
                "action": "confirm" | "modify" | "ask" | "cancel",
                "modifications": [...],
                "summary": "对修改指令的中文描述"
            }
        """
        try:
            # 构建上下文信息
            table = mapping_result.get("comparison_table", [])
            alerts = mapping_result.get("alerts", [])
            
            # 构建商品列表供LLM参考
            product_list = []
            for row in table:
                product_list.append({
                    "序号": row["seq"],
                    "客户商品名称": row["customer_product_name"],
                    "规格": row["customer_spec"],
                    "客户单位": row["customer_unit"],
                    "客户数量": row["customer_quantity"],
                    "华鼎SKU编码": row["sku_code"],
                    "SKU名称": row["sku_name"],
                    "华鼎数量": row["huading_quantity"],
                    "华鼎单位": row["huading_unit"],
                    "单位类型": row["unit_type"]
                })
            
            # 告警信息
            alert_list = [{"类型": a["type"], "行号": a["row"], "消息": a["message"]} for a in alerts]
            
            # 简单关键词匹配解析（实际使用时可接入LLM）
            import re
            
            user_msg = user_message.strip()
            
            # 确认通过
            confirm_keywords = ["没问题", "确认", "通过", "生成吧", "可以", "好的", "确认生成"]
            if any(kw in user_msg for kw in confirm_keywords):
                return {
                    "action": "confirm",
                    "modifications": [],
                    "summary": "用户确认通过"
                }
            
            # 取消
            cancel_keywords = ["取消", "算了", "不要了"]
            if any(kw in user_msg for kw in cancel_keywords):
                return {
                    "action": "cancel",
                    "modifications": [],
                    "summary": "用户取消操作"
                }
            
            # 定向修改 - 尝试正则匹配
            modifications = []
            
            # "第X行Y应该是Z" 或 "第X行Y改成Z" 或 "把第X行Y改成Z"
            row_pattern = r'第(\d+)行'
            field_pattern = r'(SKU|系统SKU|华鼎单位|单位类型|商品名称|数量|规格)'
            value_pattern = r'(?:应该是|改成|是)([^(|\n]+)'
            
            # 查找所有行号
            row_matches = re.findall(row_pattern, user_msg)
            if row_matches:
                for row_num in row_matches:
                    mod = {"row": int(row_num), "field": "", "old_value": "", "new_value": ""}
                    
                    # 尝试匹配字段
                    field_match = re.search(field_pattern, user_msg)
                    if field_match:
                        field = field_match.group(1)
                        # 字段名标准化（使用新字段名）
                        field_map = {
                            "SKU": "sku_code",
                            "华鼎SKU编码": "sku_code",
                            "SKU名称": "sku_name",
                            "华鼎单位": "huading_unit",
                            "单位类型": "unit_type",
                            "客户商品名称": "customer_product_name",
                            "数量": "quantity",
                            "规格": "customer_spec"
                        }
                        mod["field"] = field_map.get(field, field)
                    
                    # 尝试匹配新值
                    value_match = re.search(value_pattern, user_msg)
                    if value_match:
                        mod["new_value"] = value_match.group(1).strip()
                    
                    if mod["field"] and mod["new_value"]:
                        modifications.append(mod)
            
            # 全局修改 - "所有X改成Y"
            all_pattern = r'所有(华鼎单位|单位类型|SKU|数量)(?:都)?改成(.+)'
            all_match = re.search(all_pattern, user_msg)
            if all_match:
                field = all_match.group(1)
                value = all_match.group(2).strip()
                field_map = {
                    "华鼎单位": "huading_unit",
                    "单位类型": "unit_type",
                    "SKU": "sku_code",
                    "华鼎SKU编码": "sku_code",
                    "数量": "quantity"
                }
                modifications.append({
                    "row": "all",
                    "field": field_map.get(field, field),
                    "old_value": "",
                    "new_value": value
                })
            
            if modifications:
                return {
                    "action": "modify",
                    "modifications": modifications,
                    "summary": f"检测到{len(modifications)}项修改"
                }
            
            # 无法解析，默认作为提问处理
            return {
                "action": "ask",
                "modifications": [],
                "summary": "无法理解用户反馈，请重述",
                "question": user_msg
            }
            
        except Exception as e:
            print(f"解析用户反馈失败: {e}")
            return {
                "action": "ask",
                "modifications": [],
                "summary": f"解析失败: {str(e)}"
            }
    
    def _apply_modifications(self, mapping_result: Dict[str, Any], modifications: list,
                            sku_results: list, owner_code: str) -> Dict[str, Any]:
        """
        应用修改到映射结果，并重新生成模板
        
        Args:
            mapping_result: 当前映射结果
            modifications: 修改指令列表
            sku_results: SKU匹配结果
            owner_code: 货主ID
            
        Returns:
            更新后的 mapping_result
        """
        comparison_table = mapping_result["comparison_table"]
        
        for mod in modifications:
            row = mod["row"]
            field = mod["field"]
            new_value = mod["new_value"]
            
            if row == "all":
                # 全局修改
                for item in comparison_table:
                    if field == "unit_type" and item["matched"]:
                        # 单位类型修改需要重新获取华鼎单位
                        sku_code = item["sku_code"]
                        unit_info = object.__getattribute__(self, '_get_sku_unit_info')(sku_code, owner_code)
                        item["unit_type"] = new_value
                        item["huading_unit"] = unit_info["unit"]
                    elif field in item:
                        item[field] = new_value
            else:
                # 定向修改
                for item in comparison_table:
                    if item["seq"] == row:
                        if field == "unit_type" and item["matched"]:
                            # 单位类型修改需要重新获取华鼎单位
                            sku_code = item["sku_code"]
                            unit_info = object.__getattribute__(self, '_get_sku_unit_info')(sku_code, owner_code)
                            item["unit_type"] = new_value
                            item["huading_unit"] = unit_info["unit"]
                        elif field in item:
                            item[field] = new_value
                        break
        
        # 更新mapping_result
        mapping_result["comparison_table"] = comparison_table
        
        # 重新检查告警
        alerts = []
        for row in comparison_table:
            if not row["matched"]:
                alerts.append({
                    "type": "unmatched",
                    "row": row["seq"],
                    "severity": "🔴",
                    "message": f"第{row['seq']}行商品未匹配到系统SKU"
                })
            elif not row["huading_unit"]:
                alerts.append({
                    "type": "empty_field",
                    "row": row["seq"],
                    "severity": "🔴",
                    "message": f"第{row['seq']}行华鼎单位为空"
                })
        
        mapping_result["alerts"] = alerts
        mapping_result["summary"]["alert_count"] = len(alerts)
        mapping_result["summary"]["has_critical_alerts"] = any(a["type"] in ["unmatched", "empty_field"] for a in alerts)
        
        return mapping_result
    
    def _generate_template(self, order_data: Dict, store_info: Optional[Dict], 
                          sku_results: list, output_file: str):
        """生成华鼎模板Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "omsWare"
        
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        
        column_widths = {1: 6, 2: 16, 3: 12, 4: 8, 5: 10, 6: 18, 7: 16, 8: 10, 9: 10,
            10: 10, 11: 8, 12: 10, 13: 12, 14: 8, 15: 8, 16: 10, 17: 8, 18: 10,
            19: 10, 20: 15, 21: 12, 22: 25, 23: 12, 24: 15, 25: 18, 26: 8, 27: 8,
            28: 10, 29: 15, 30: 40, 31: 12}
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[1].height = 35
        
        cell_alignment = Alignment(horizontal="center", vertical="center")
        cell_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # ⚠️ 仓库编码必须有值，不能为空
        warehouse_code = ""
        if store_info:
            warehouse_code = store_info.get("warehouse_code", "")
            if not warehouse_code and store_info.get("warehouse_name"):
                warehouse_code = object.__getattribute__(self, '_get_warehouse_code')(store_info.get("warehouse_name", ""))
        
        # 如果仓库编码仍然为空，报错（这是严重错误）
        if not warehouse_code:
            raise ValueError(f"仓库编码不能为空！store_info={store_info}")
        
        for row_idx, item in enumerate(sku_results, start=2):
            row_data = {
                "序号": 1,
                "门店编号": store_info["store_code"] if store_info else "",
                "门店三方编码": "",
                "仓库编码": warehouse_code,
                "加急程度\n（0：普通，1：加急）": self.DEFAULT_VALUES["加急程度"],
                "商品SKU编号": item["sku_code"],
                "商品三方SPEC编号": "",
                "单位类型": item["unit_type"],
                "出库数量": item["quantity"],
                "指定库存状态": self.DEFAULT_VALUES["指定库存状态"],
                "出库类型": self.DEFAULT_VALUES["出库类型"],
                "配送方式": self.DEFAULT_VALUES["配送方式"],
                "指定车型（专配）": "",
                "是否垫付": self.DEFAULT_VALUES["是否垫付"],
                "付款方式": "",
                "快递公司": "",
                "单价": "",
                "总金额": "",
                "是否制定批次": self.DEFAULT_VALUES["是否制定批次"],
                "批次号": "",
                "生产日期": "",
                "备注": item["remark"],
                "生产厂家编号": "",
                "门店收货地址编码": "",
                "三方单号": order_data["order_no"],
                "业务模式": "",
                "业务类型": "",
                "收货人": store_info["contact_person"] if store_info else "",
                "联系电话": store_info["phone"] if store_info else "",
                "收货地址": store_info["address"] if store_info else "",
                "C端快递公司": ""
            }
            
            for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
                value = row_data.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = cell_border
        
        wb.save(output_file)

    def _generate_multi_store_template(self, order_data: Dict, all_store_results: list, output_file: str):
        """
        生成多门店合并模板（所有门店写入同一个Sheet）

        格式：每个门店之间用空行分隔，门店名作为分隔标识
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "omsWare"

        # 样式
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")
        cell_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        store_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        store_font = Font(bold=True, color="000000", size=10)

        # 写表头
        for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        column_widths = {1:6, 2:16, 3:12, 4:8, 5:10, 6:18, 7:16, 8:10, 9:10,
                         10:10, 11:8, 12:10, 13:12, 14:8, 15:8, 16:10, 17:8, 18:10,
                         19:10, 20:15, 21:12, 22:25, 23:12, 24:15, 25:18, 26:8, 27:8,
                         28:10, 29:15, 30:40, 31:12}
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 35

        row_idx = 2
        
        for store_idx, store_result in enumerate(all_store_results, start=1):
            si = store_result["store_info"]
            sku_results = store_result["sku_results"]

            # 商品明细行（序号=门店序号，同门店商品序号相同，从1开始）
            for sku in sku_results:
                row_data = {
                    "序号": store_idx,  # 序号=门店序号（同一门店的商品序号相同，从1开始）
                    "门店编号": si.get("store_code", ""),
                    "门店三方编码": "",
                    "仓库编码": si.get("warehouse_code", ""),
                    "加急程度\n（0：普通，1：加急）": self.DEFAULT_VALUES["加急程度"],
                    "商品SKU编号": sku.get("sku_code", ""),
                    "商品三方SPEC编号": "",
                    "单位类型": sku.get("unit_type", "大单位"),
                    "出库数量": sku.get("quantity", 1),
                    "指定库存状态": self.DEFAULT_VALUES["指定库存状态"],
                    "出库类型": self.DEFAULT_VALUES["出库类型"],
                    "配送方式": self.DEFAULT_VALUES["配送方式"],
                    "指定车型（专配）": "",
                    "是否垫付": self.DEFAULT_VALUES["是否垫付"],
                    "付款方式": "",
                    "快递公司": "",
                    "单价": "",
                    "总金额": "",
                    "是否制定批次": self.DEFAULT_VALUES["是否制定批次"],
                    "批次号": "",
                    "生产日期": "",
                    "备注": sku.get("remark", ""),
                    "生产厂家编号": "",
                    "门店收货地址编码": "",
                    "三方单号": order_data.get("order_no", ""),
                    "业务模式": "",
                    "业务类型": "",
                    "收货人": si.get("contact_person", ""),
                    "联系电话": si.get("phone", ""),
                    "收货地址": si.get("address", ""),
                    "C端快递公司": "",
                }
                for col_idx, field_name in enumerate(self.HUADING_FIELDS, start=1):
                    value = row_data.get(field_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = cell_border
                row_idx += 1

        wb.save(output_file)

    def _generate_mapping_comparison_multi(self, order_data: Dict, all_store_results: list) -> Dict:
        """
        为多门店订单生成汇总映射对照表
        
        ⚠️ 映射对照表必须使用以下9列格式：
        - 订单商品名称：订单原始商品名称
        - 订单商品规格：订单原始规格
        - 订单数量：订单原始数量
        - 订单单位：订单原始单位
        - 匹配SKU编码：数据库匹配到的SKU编码
        - SKU名称：数据库SKU名称
        - 数量：出库数量
        - 单位类型：大单位/小单位
        - 匹配单位：对应出库单位（箱/件等）
        """
        all_mappings = []
        total_alert_count = 0
        total_matched_count = 0

        for store_result in all_store_results:
            si = store_result["store_info"]
            sku_results = store_result["sku_results"]
            store_items = store_result["items"]
            store_name = store_result["store_name"]

            for idx, sku in enumerate(sku_results):
                confidence = sku.get("confidence", 1.0)
                is_alert = confidence < 0.8
                if is_alert:
                    total_alert_count += 1
                if confidence >= 0.8:
                    total_matched_count += 1

                # 获取订单原始数据
                order_item = store_items[idx] if idx < len(store_items) else {}
                
                all_mappings.append({
                    # 订单原始信息
                    "订单商品名称": sku.get("original_product_name", order_item.get("product_name", "")),
                    "订单商品规格": order_item.get("spec", ""),
                    "订单数量": order_item.get("quantity", sku.get("quantity", 0)),
                    "订单单位": order_item.get("unit", "件"),
                    # 匹配结果
                    "匹配SKU编码": sku.get("sku_code", ""),
                    "SKU名称": sku.get("sku_name", ""),
                    "数量": sku.get("quantity", 0),
                    "单位类型": sku.get("unit_type", "大单位"),
                    "匹配单位": sku.get("unit", "件"),
                    # 辅助字段（用于调试）
                    "store": store_name,
                    "seq": idx + 1,
                    "confidence": confidence,
                    "is_alert": is_alert,
                    "alert_reason": "置信度<80%" if is_alert else "",
                })

            for item in store_result["unmatched_items"]:
                all_mappings.append({
                    # 订单原始信息
                    "订单商品名称": item.get("product_name", ""),
                    "订单商品规格": item.get("spec", ""),
                    "订单数量": item.get("quantity", 0),
                    "订单单位": item.get("unit", "件"),
                    # 匹配结果
                    "匹配SKU编码": "",
                    "SKU名称": "",
                    "数量": "",
                    "单位类型": "",
                    "匹配单位": "",
                    # 辅助字段
                    "store": store_name,
                    "seq": "❌",
                    "confidence": 0,
                    "is_alert": True,
                    "alert_reason": "SKU未匹配",
                })

        return {
            "mappings": all_mappings,
            "summary": {
                "total_items": len(all_mappings),
                "matched_count": total_matched_count,
                "alert_count": total_alert_count,
                "unmatched_count": sum(len(r["unmatched_items"]) for r in all_store_results),
                "store_count": len(all_store_results),
            }
        }


# 全局配置
_global_config = {
    "shipper_id": None,
    "db_config": None,
    "output_dir": "./output"
}


def configure(shipper_id: str, db_config: Dict, output_dir: Optional[str] = None):
    """配置全局参数"""
    _global_config["shipper_id"] = shipper_id
    _global_config["db_config"] = db_config
    _global_config["output_dir"] = output_dir or "./output"


def convert_order_to_huading(order_file: str, output_file: str = None) -> Dict[str, Any]:
    """将客户订单转换为华鼎出库单模板"""
    if not _global_config["shipper_id"] or not _global_config["db_config"]:
        return {
            "success": False,
            "message": "请先调用 configure() 配置全局参数"
        }
    
    skill = OrderToHuadingTemplate(
        shipper_id=_global_config["shipper_id"],
        db_config=_global_config["db_config"],
        output_dir=_global_config["output_dir"]
    )
    
    return skill.execute(order_file, output_file)


if __name__ == "__main__":
    print("请先调用 configure() 配置参数")