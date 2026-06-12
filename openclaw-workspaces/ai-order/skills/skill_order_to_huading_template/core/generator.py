"""
core/generator.py — 模板生成核心模块

职责：
1. 生成华鼎 31 字段出库单 Excel（单门店 + 多门店合并）
2. 格式化映射对照表为文本（飞书展示用）
3. 格式化成功返回的用户消息
4. 文件发送准备（get_file_info, prepare_file_for_send）
5. 文件下载 URL 生成

提取来源：
- __init__.py: _generate_template, _generate_multi_store_template,
  _format_comparison_table_text, _format_success_message,
  get_file_info, prepare_file_for_send, _get_download_url
"""
import os
from typing import Dict, Any, Optional, List

from config import _get_huading_fields


# ---------------------------------------------------------------------------
# 华鼎模板字段 + 默认值
# ---------------------------------------------------------------------------

HUADING_FIELDS = list(_get_huading_fields())

DEFAULT_VALUES = {
    "加急程度": 0,
    "指定库存状态": "正常",
    "出库类型": 201,
    "配送方式": "共配",
    "是否垫付": "否",
    "是否制定批次": "否"
}

COLUMN_WIDTHS = {
    1: 6, 2: 16, 3: 12, 4: 8, 5: 10, 6: 18, 7: 16, 8: 10, 9: 10,
    10: 10, 11: 8, 12: 10, 13: 12, 14: 8, 15: 8, 16: 10, 17: 8, 18: 10,
    19: 10, 20: 15, 21: 12, 22: 25, 23: 12, 24: 15, 25: 18, 26: 8, 27: 8,
    28: 10, 29: 15, 30: 40, 31: 12
}


# ---------------------------------------------------------------------------
# 文件下载 URL
# ---------------------------------------------------------------------------

AWS_PUBLIC_IP = os.getenv("AWS_PUBLIC_IP", "")
AWS_FILE_PORT = int(os.getenv("AWS_FILE_PORT", "0"))


def get_download_url(output_file: str) -> str:
    """生成文件下载 URL（基于 AWS 公网 IP）"""
    if not output_file or not AWS_PUBLIC_IP or not AWS_FILE_PORT:
        return ""
    filename = os.path.basename(output_file)
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    return f"http://{AWS_PUBLIC_IP}:{AWS_FILE_PORT}/{encoded_filename}"


# ---------------------------------------------------------------------------
# 文件信息
# ---------------------------------------------------------------------------

def get_file_info(file_path: str) -> Dict[str, Any]:
    """获取文件信息（用于发送文件给用户）"""
    if not os.path.exists(file_path):
        return {
            "exists": False,
            "file_name": os.path.basename(file_path),
            "error": "文件不存在"
        }
    file_size = os.path.getsize(file_path)
    if file_size > 1024 * 1024:
        size_str = f"{file_size / (1024 * 1024):.1f} MB"
    elif file_size > 1024:
        size_str = f"{file_size / 1024:.0f} KB"
    else:
        size_str = f"{file_size} B"
    return {
        "exists": True,
        "file_name": os.path.basename(file_path),
        "file_size": file_size,
        "file_size_str": size_str,
        "path": file_path,
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }


def prepare_file_for_send(file_path: str) -> Dict[str, Any]:
    """准备文件发送所需信息"""
    file_info = get_file_info(file_path)
    if not file_info["exists"]:
        return {"success": False, "error": file_info["error"]}
    return {
        "success": True,
        "file_path": file_path,
        "file_name": file_info["file_name"],
        "file_size_str": file_info["file_size_str"],
        "can_send": True,
        "for_message": {
            "media": file_path,
            "file_name": file_info["file_name"]
        }
    }


# ---------------------------------------------------------------------------
# 成功消息格式化
# ---------------------------------------------------------------------------

def format_success_message(store_names: str, item_count: int,
                           matched_count: int, unmatched_count: int,
                           file_name: str, has_issues: bool = False) -> str:
    """格式化成功返回的用户消息"""
    lines = []
    lines.append(f"✅ 订单处理完成")
    lines.append(f"")
    lines.append(f"📦 订单信息")
    lines.append(f"   门店：{store_names}")
    lines.append(f"   商品数：{item_count}条")
    lines.append(f"")
    lines.append(f"📊 匹配结果")
    lines.append(f"   ✅ 匹配成功：{matched_count}条")
    if unmatched_count > 0:
        lines.append(f"   ⚠️ 未匹配：{unmatched_count}条（需人工确认）")
    else:
        lines.append(f"   ⬜ 未匹配：0条")
    lines.append(f"")
    lines.append(f"📁 输出文件")
    lines.append(f"   {file_name}")
    lines.append(f"")
    if has_issues:
        lines.append(f"⚠️ 温馨提示：有{unmatched_count}条商品未匹配成功，请检查后确认")
    else:
        lines.append(f"💡 请下载文件查看详情，确认无误后提交审核")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 映射对照表文本格式化
# ---------------------------------------------------------------------------

def format_comparison_table_text(mapping_result: Dict[str, Any], order_info: Dict = None) -> str:
    """格式化映射对照表为文本（飞书展示用）"""
    table = mapping_result["comparison_table"]
    alerts = mapping_result["alerts"]
    summary = mapping_result["summary"]

    lines = []

    if order_info:
        lines.append("━" * 50)
        lines.append("订单信息")
        lines.append(f"  订单号: {order_info.get('order_no', '-')}")
        lines.append(f"  门店: {order_info.get('store_name', '-')}")
        lines.append(f"  货主: {order_info.get('owner_name', '-')}")
        lines.append(f"  仓库: {order_info.get('warehouse_name', '-')}")
        lines.append("━" * 50)
        lines.append("")

    if alerts:
        lines.append("━" * 50)
        lines.append(f"告警 ({len(alerts)} 项)")
        for a in alerts:
            lines.append(f"  {a['severity']} {a['message']}")
        lines.append("━" * 50)
        lines.append("")

    status = "✅" if not summary["has_critical_alerts"] else "⚠️ 需要检查"
    lines.append(f"商品映射对照表（{summary['total_items']}条）{status}")
    lines.append(f"  已匹配: {summary['matched_count']} | 未匹配: {summary['unmatched_count']} | 告警: {summary['alert_count']}")
    lines.append("")

    header = "| # | 客户商品名称 | 规格 | 单位 | 数量 | → | 华鼎SKU编码 | SKU名称 | 数量 | 单位 | 单位类型 | 状态|"
    separator = "|---|-------------|------|------|------|---|-------------|---------|------|------|----------|------|"
    lines.append(header)
    lines.append(separator)

    alert_rows = {a["row"]: a for a in alerts}

    for row in table:
        seq = row["seq"]
        if not row["matched"]:
            status_icon = "🔴"
        elif row["confidence"] < 0.8:
            status_icon = "⚠️"
        else:
            status_icon = "✅"

        sku_code = row["sku_code"] or "-"
        sku_name = row["sku_name"] or "-"
        huading_quantity = row["huading_quantity"] if row["huading_quantity"] != "" else "-"
        huading_unit = row["huading_unit"] or "-"
        unit_type = row["unit_type"] or "-"

        line = f"| {seq} | {row['customer_product_name']} | {row['customer_spec']} | {row['customer_unit']} | {row['customer_quantity']} | → | {sku_code} | {sku_name} | {huading_quantity} | {huading_unit} | {unit_type} | {status_icon}|"
        lines.append(line)

        if row.get("need_confirm") and row.get("candidates"):
            lines.append(f"  ↳ ⚠️ 第{seq}行有多个同名SKU，请选择：")
            for ci, c in enumerate(row["candidates"], 1):
                c_sku = c.get("sku_code", "")
                c_name = c.get("sku_name", "")
                c_unit = c.get("unit", "")
                c_type = c.get("unit_type", "")
                c_spec = c.get("product_spec", "")
                selected = " ← 当前选中" if c_sku == sku_code else ""
                lines.append(f"    {ci}. {c_sku} | {c_name} | {c_unit} | {c_type} | 规格:{c_spec}{selected}")

    lines.append("")
    lines.append("请输入指令：")
    lines.append("  - 确认 / 没问题 / 通过 → 生成模板")
    lines.append("  - 修改第X行 → 修改指定行")
    lines.append("  - 取消 → 放弃订单")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# TemplateGenerator 类
# ---------------------------------------------------------------------------

class TemplateGenerator:
    """
    华鼎出库单模板生成器

    用法：
        gen = TemplateGenerator(warehouse_code_map)
        gen.generate_single(order_data, store_info, sku_results, output_file)
        gen.generate_multi(order_data, all_store_results, output_file)
    """

    def __init__(self, warehouse_code_map: Dict[str, str] = None):
        self.warehouse_code_map = warehouse_code_map or {}

    def get_warehouse_code(self, warehouse_name: str) -> str:
        """获取仓库编码"""
        if not warehouse_name:
            raise ValueError("门店的仓库名称为空，请检查门店数据")
        code = self.warehouse_code_map.get(warehouse_name)
        if not code:
            available = ", ".join(self.warehouse_code_map.keys())
            raise ValueError(
                f"仓库'{warehouse_name}'在仓库编码映射表中未找到！\n"
                f"当前可用的仓库：{available}"
            )
        return code

    def generate_single(self, order_data: Dict, store_info: Optional[Dict],
                        sku_results: list, output_file: str):
        """生成单门店华鼎模板 Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "omsWare"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

        for col_idx, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 35

        cell_alignment = Alignment(horizontal="center", vertical="center")
        cell_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        warehouse_code = ""
        if store_info:
            warehouse_code = store_info.get("warehouse_code", "")
            if not warehouse_code and store_info.get("warehouse_name"):
                warehouse_code = self.get_warehouse_code(store_info.get("warehouse_name", ""))

        if not warehouse_code:
            raise ValueError(f"仓库编码不能为空！store_info={store_info}")

        for row_idx, item in enumerate(sku_results, start=2):
            row_data = {
                "序号": 1,
                "门店编号": store_info["store_code"] if store_info else "",
                "门店三方编码": "",
                "仓库编码": warehouse_code,
                "加急程度\n（0：普通，1：加急）": DEFAULT_VALUES["加急程度"],
                "商品SKU编号": item["sku_code"],
                "商品三方SPEC编号": "",
                "单位类型": item["unit_type"],
                "出库数量": item["quantity"],
                "指定库存状态": DEFAULT_VALUES["指定库存状态"],
                "出库类型": DEFAULT_VALUES["出库类型"],
                "配送方式": DEFAULT_VALUES["配送方式"],
                "指定车型（专配）": "",
                "是否垫付": DEFAULT_VALUES["是否垫付"],
                "付款方式": "",
                "快递公司": "",
                "单价": "",
                "总金额": "",
                "是否制定批次": DEFAULT_VALUES["是否制定批次"],
                "批次号": "",
                "生产日期": "",
                "备注": item["remark"],
                "生产厂家编号": "",
                "门店收货地址编码": "",
                "三方单号": order_data["order_no"],
                "业务模式": "",
                "业务类型": "",
                "收货人": store_info["contact_person"] if store_info else "",
                "联系电话": store_info["phone"] if store_info else "",
                "收货地址": store_info["address"] if store_info else "",
                "C端快递公司": ""
            }
            for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
                value = row_data.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = cell_border

        wb.save(output_file)

    def generate_multi(self, order_data: Dict, all_store_results: list, output_file: str):
        """生成多门店合并模板（所有门店写入同一个 Sheet）"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "omsWare"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")
        cell_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # 多门店仓库编码强校验
        missing_warehouse = []
        for store_result in all_store_results:
            si = store_result["store_info"]
            if not si.get("warehouse_code", ""):
                missing_warehouse.append(si.get("store_name", store_result.get("store_name", "未知门店")))
        if missing_warehouse:
            raise ValueError(
                f"以下门店缺少仓库编码，无法生成模板：{', '.join(missing_warehouse)}。"
                f"请检查仓库配置或确认门店匹配结果。"
            )

        for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        for col_idx, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 35

        row_idx = 2

        for store_idx, store_result in enumerate(all_store_results, start=1):
            si = store_result["store_info"]
            sku_results = store_result["sku_results"]

            for sku in sku_results:
                row_data = {
                    "序号": store_idx,
                    "门店编号": si.get("store_code", ""),
                    "门店三方编码": "",
                    "仓库编码": si.get("warehouse_code", ""),
                    "加急程度\n（0：普通，1：加急）": DEFAULT_VALUES["加急程度"],
                    "商品SKU编号": sku.get("sku_code", ""),
                    "商品三方SPEC编号": "",
                    "单位类型": sku.get("unit_type", ""),
                    "出库数量": sku.get("quantity", 1),
                    "指定库存状态": DEFAULT_VALUES["指定库存状态"],
                    "出库类型": DEFAULT_VALUES["出库类型"],
                    "配送方式": DEFAULT_VALUES["配送方式"],
                    "指定车型（专配）": "",
                    "是否垫付": DEFAULT_VALUES["是否垫付"],
                    "付款方式": "",
                    "快递公司": "",
                    "单价": "",
                    "总金额": "",
                    "是否制定批次": DEFAULT_VALUES["是否制定批次"],
                    "批次号": "",
                    "生产日期": "",
                    "备注": sku.get("remark", ""),
                    "生产厂家编号": "",
                    "门店收货地址编码": "",
                    "三方单号": order_data.get("order_no", ""),
                    "业务模式": "",
                    "业务类型": "",
                    "收货人": si.get("contact_person", ""),
                    "联系电话": si.get("phone", ""),
                    "收货地址": si.get("address", ""),
                    "C端快递公司": "",
                }
                for col_idx, field_name in enumerate(HUADING_FIELDS, start=1):
                    value = row_data.get(field_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = cell_border
                row_idx += 1

        wb.save(output_file)
