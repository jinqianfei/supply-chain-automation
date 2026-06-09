"""
模板生成工具 - 生成华鼎31字段出库单模板

华鼎模板格式固定，不做规则配置
"""
import os
import sys
import json
from datetime import datetime
from typing import Optional, Dict, Any, List
import yaml
from db.connection import get_connection

# 允许独立运行：加入 skill 根目录
_SKILL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SKILL_ROOT not in sys.path:
    sys.path.insert(0, _SKILL_ROOT)

from config import get_huading_fields  # v5.11.2 统一从 yaml 读
from db.table_names import SKU_TABLE, WAREHOUSE_TABLE  # v5.11.2 表名常量


# 华鼎31字段（v5.11.2 改为从 yaml 读，保持同名变量向后兼容）
HUADING_FIELDS = get_huading_fields()


def load_defaults() -> dict:
    """加载模板默认值配置"""
    config_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "config", "template_defaults.yaml"
    )
    if os.path.exists(config_path):
        with open(config_path, "r") as f:
            data = yaml.safe_load(f)
            return data.get("huading_template", {}).get("defaults", {})
    return {}


def get_warehouse_code(warehouse_name: str, db_config: Optional[dict] = None) -> str:
    """获取仓库编码"""
    if not warehouse_name:
        raise ValueError("仓库名称为空")
    
    conn = get_connection(db_config)
    cur = conn.cursor()
    
    cur.execute("""
        SELECT warehouse_code FROM {WAREHOUSE_TABLE}
        WHERE warehouse_name = %s
    """, (warehouse_name,))
    
    row = cur.fetchone()
    cur.close()
    conn.close()
    
    if not row:
        raise ValueError(f"仓库'{warehouse_name}'未在仓库编码映射表中找到")
    
    return row[0]


def generate(store_info: dict, sku_mappings: list, order_data: dict,
             output_dir: str = "./output", order_no: str = "",
             db_config: Optional[dict] = None,
             store_seq: int = 1) -> str:
    """
    生成华鼎31字段出库单模板
    
    Args:
        store_info: 门店信息（含仓库编码、收货人信息）
        sku_mappings: SKU映射结果列表
        order_data: 订单信息（统一JSON）
        output_dir: 输出目录
        order_no: 订单号
        db_config: 数据库配置
        store_seq: 门店序号（同一门店的商品序号相同，从1开始）
    
    Returns:
        .xlsx 文件路径
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    defaults = load_defaults()
    
    # 生成订单号
    now = datetime.now()
    order_no = order_no or f"DH-O-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')}"
    
    # 构建输出路径
    safe_order_no = order_no.replace("/", "-").replace("\\", "-")
    output_file = os.path.join(output_dir, f"华鼎出库单_{safe_order_no}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    
    # 获取仓库编码
    warehouse_code = store_info.get("warehouse_code", "") or get_warehouse_code(
        store_info.get("warehouse_name", ""), db_config
    )
    
    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "omsWare"
    
    # 样式
    header_font = Font(bold=True, size=10)
    header_fill = openpyxl.styles.PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 写表头
    for col, field in enumerate(HUADING_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # 写数据行
    for idx, sku in enumerate(sku_mappings):
        row_num = idx + 2
        seq = store_seq  # 序号=门店序号（同一门店的商品序号相同，从1开始）
        
        ws.cell(row=row_num, column=1, value=seq)
        ws.cell(row=row_num, column=2, value=store_info.get("store_code", ""))
        ws.cell(row=row_num, column=3, value=defaults.get("门店三方编码", ""))
        ws.cell(row=row_num, column=4, value=warehouse_code)
        ws.cell(row=row_num, column=5, value=defaults.get("加急程度", 0))
        ws.cell(row=row_num, column=6, value=sku.get("sku_code", ""))
        ws.cell(row=row_num, column=7, value=defaults.get("商品三方SPEC编号", ""))
        ws.cell(row=row_num, column=8, value=sku.get("unit_type", "大单位"))
        ws.cell(row=row_num, column=9, value=sku.get("quantity", 1))
        ws.cell(row=row_num, column=10, value=defaults.get("指定库存状态", "正常"))
        ws.cell(row=row_num, column=11, value=defaults.get("出库类型", "201"))
        ws.cell(row=row_num, column=12, value=defaults.get("配送方式", "共配"))
        ws.cell(row=row_num, column=13, value=defaults.get("指定车型（专配）", ""))
        ws.cell(row=row_num, column=14, value=defaults.get("是否垫付", "否"))
        ws.cell(row=row_num, column=15, value=defaults.get("付款方式", ""))
        ws.cell(row=row_num, column=16, value=defaults.get("快递公司", ""))
        ws.cell(row=row_num, column=17, value=defaults.get("单价", ""))
        ws.cell(row=row_num, column=18, value=defaults.get("总金额", ""))
        ws.cell(row=row_num, column=19, value=defaults.get("是否制定批次", "否"))
        ws.cell(row=row_num, column=20, value=defaults.get("批次号", ""))
        ws.cell(row=row_num, column=21, value=defaults.get("生产日期", ""))
        ws.cell(row=row_num, column=22, value=sku.get("remark", order_data.get("extra_notes", "")))
        ws.cell(row=row_num, column=23, value=defaults.get("生产厂家编号", ""))
        ws.cell(row=row_num, column=24, value=defaults.get("门店收货地址编码", ""))
        ws.cell(row=row_num, column=25, value=order_no)
        ws.cell(row=row_num, column=26, value=defaults.get("业务模式", ""))
        ws.cell(row=row_num, column=27, value=defaults.get("业务类型", ""))
        ws.cell(row=row_num, column=28, value=store_info.get("contact_person", ""))
        ws.cell(row=row_num, column=29, value=store_info.get("phone", ""))
        ws.cell(row=row_num, column=30, value=store_info.get("address", ""))
        ws.cell(row=row_num, column=31, value=defaults.get("C端快递公司", ""))
        
        for col in range(1, 32):
            ws.cell(row=row_num, column=col).border = thin_border
            ws.cell(row=row_num, column=col).alignment = center_align
    
    # 设置列宽
    col_widths = [6, 16, 12, 10, 10, 20, 16, 10, 10, 10, 10, 10, 14, 10, 10, 10, 10, 10, 12, 10, 12, 10, 12, 16, 26, 10, 10, 12, 16, 40, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    wb.save(output_file)
    return output_file


def generate_from_template(template_path: str, store_info: dict,
                           sku_mappings: list, order_data: dict,
                           output_path: str, order_no: str = "") -> str:
    """
    基于已有模板文件生成出库单（保留原有公式和格式）
    
    当用户提供了模板文件时使用此方法
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    import shutil
    
    shutil.copy2(template_path, output_path)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb['omsWare']
    
    defaults = load_defaults()
    now = datetime.now()
    order_no = order_no or f"DH-O-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')}"
    warehouse_code = store_info.get("warehouse_code", "")
    
    store1_code = store_info.get("store_code", "")
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for idx, sku in enumerate(sku_mappings):
        row_num = idx + 2
        
        # 单位类型（H）
        ws.cell(row=row_num, column=8, value=sku.get("unit_type", "大单位"))
        # 是否垫付（N）
        ws.cell(row=row_num, column=14, value=defaults.get("是否垫付", "否"))
        # 收货人（Z）
        ws.cell(row=row_num, column=26, value=store_info.get("contact_person", ""))
        # 联系电话（AA）
        ws.cell(row=row_num, column=27, value=store_info.get("phone", ""))
        # 收货地址（AB）
        ws.cell(row=row_num, column=28, value=store_info.get("address", ""))
        # 备注（V）- 保留原公式不变
        
        for col in range(1, 30):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = center_align
    
    wb.save(output_path)
    return output_path
