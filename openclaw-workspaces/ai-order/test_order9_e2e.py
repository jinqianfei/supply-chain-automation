#!/usr/bin/env python3
"""
端到端跑 9 号订单（天津仓 2 店 11 项）
"""
import json
import os
import sys
import re
import time

# 加载 .env
from pathlib import Path
env_file = Path("/Users/jinqianfei/openclaw-workspaces/ai-order/.env")
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, v = line.split("=", 1)
            os.environ[k] = v

# 添加 skill 路径
SKILL_DIR = "/Users/jinqianfei/openclaw-workspaces/ai-order/skills/skill_order_to_huading_template"
sys.path.insert(0, SKILL_DIR)

from __init__ import OrderToHuadingTemplate  # noqa: E402

DB_CONFIG = {
    "host": os.environ.get("DB_HOST", "agenthub-db.cjys0msc4x8s.ap-southeast-1.rds.amazonaws.com"),
    "port": int(os.environ.get("DB_PORT", 5432)),
    "database": os.environ.get("DB_NAME", "neo"),
    "user": os.environ.get("DB_USER", "agenthub"),
    "password": os.environ.get("DB_PASSWORD", ""),
}

INPUT_XLSX = "/Users/jinqianfei/openclaw-workspaces/ai-order/data/test_orders/天津仓_2店11项.xlsx"
OUTPUT_XLSX = "/Users/jinqianfei/openclaw-workspaces/ai-order/output/test_天津仓_2店11项_端到端.xlsx"

print("=" * 80)
print("端到端跑 9 号订单（天津仓 2 店 11 项）")
print("=" * 80)
print(f"DB: host={DB_CONFIG['host']}, db={DB_CONFIG['database']}, user={DB_CONFIG['user']}")
print(f"Input:  {INPUT_XLSX}")
print(f"Output: {OUTPUT_XLSX}")
print()

skill = OrderToHuadingTemplate(
    db_config=DB_CONFIG,
    output_dir="/Users/jinqianfei/openclaw-workspaces/ai-order/output",
)
print(f"Skill version: {OrderToHuadingTemplate.VERSION}")
print()

# 直接调 execute()
t0 = time.time()
result = skill.execute(order_input=INPUT_XLSX, output_file=OUTPUT_XLSX, order_type="excel")
t1 = time.time()
print(f"\n>>> execute() 耗时: {t1 - t0:.2f}s")
print()

# ── 1. 基本信息 ──
print("=" * 80)
print("1. 基本结果")
print("=" * 80)
for k in ["success", "need_review", "need_ocr", "need_store_confirm", "need_customer_hint",
          "file_name", "order_no", "store_names", "store_count",
          "item_count", "matched_count", "unmatched_count", "has_issues",
          "extracted_from", "message"]:
    v = result.get(k)
    if isinstance(v, str) and len(v) > 200:
        v = v[:200] + "..."
    print(f"  {k}: {v}")
print()

# ── 2. all_store_results 完整内容 ──
print("=" * 80)
print("2. all_store_results 完整内容")
print("=" * 80)
asr = result.get("all_store_results", [])
print(f"  Number of stores: {len(asr)}")
for i, r in enumerate(asr, 1):
    print(f"\n  ── Store #{i}: {r.get('store_name')!r} ──")
    si = r.get("store_info") or {}
    print(f"    store_info keys: {list(si.keys())}")
    for k in ["store_code", "store_name", "owner_code", "owner_name",
              "warehouse_name", "warehouse_code", "address",
              "contact_person", "phone", "match_type", "match_method", "similarity"]:
        v = si.get(k)
        if v is not None:
            v_str = str(v)
            if len(v_str) > 80:
                v_str = v_str[:80] + "..."
            print(f"      {k}: {v_str}")
    print(f"    sku_results: {len(r.get('sku_results') or [])} items")
    for j, s in enumerate(r.get("sku_results") or [], 1):
        # 截断过长的字符串
        s_short = {}
        for k, v in s.items():
            if isinstance(v, str) and len(v) > 100:
                s_short[k] = v[:100] + "..."
            else:
                s_short[k] = v
        print(f"      #{j}: {s_short}")
    print(f"    unmatched_items: {len(r.get('unmatched_items') or [])} items")
    for j, u in enumerate(r.get("unmatched_items") or [], 1):
        u_short = {}
        for k, v in u.items():
            if isinstance(v, str) and len(v) > 100:
                u_short[k] = v[:100] + "..."
            else:
                u_short[k] = v
        print(f"      unmatched #{j}: {u_short}")
print()

# ── 3. review_data.mappings 完整内容 ──
print("=" * 80)
print("3. review_data.mappings 完整内容")
print("=" * 80)
rd = result.get("review_data") or {}
print(f"  review_data top-level keys: {list(rd.keys())}")
mappings = rd.get("mappings", [])
print(f"  mappings count: {len(mappings)}")
for i, m in enumerate(mappings, 1):
    print(f"\n  ── Mapping #{i} ──")
    for k, v in m.items():
        if isinstance(v, str) and len(v) > 100:
            v = v[:100] + "..."
        print(f"    {k}: {v}")
print()

# ── 4. 输出 xlsx 的 sheet/列/前 5 行 ──
print("=" * 80)
print("4. 输出 xlsx 内容")
print("=" * 80)
print(f"  output_file: {result.get('output_file')}")
print(f"  file exists: {os.path.exists(OUTPUT_XLSX)}")
if os.path.exists(OUTPUT_XLSX):
    import openpyxl
    wb = openpyxl.load_workbook(OUTPUT_XLSX)
    print(f"  sheet names: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  ── Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ──")
        # 打印列头
        headers = [c.value for c in ws[1]]
        print(f"    Headers: {headers}")
        # 打印前 5 行
        print("    First 5 rows:")
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True), 2):
            row_str = [str(v) if v is not None else "" for v in row]
            # 截断
            row_str = [(s[:30] + "...") if len(s) > 30 else s for s in row_str]
            print(f"      Row {r_idx}: {row_str}")
print()

# ── 5. 跟 GT 对比 ──
print("=" * 80)
print("5. 跟 GT 对比（test_set_A_history回流.json 第 9 个订单）")
print("=" * 80)
with open("/Users/jinqianfei/openclaw-workspaces/ai-order/docs/test_data/test_set_A_history回流.json") as f:
    gt_data = json.load(f)
order9_gt = gt_data[8]
print(f"  GT source: {order9_gt.get('source')}")
print()
gt = order9_gt.get("ground_truth", {})
gt_stores = order9_gt.get("stores", {})

# 拼接实际匹配结果 (按 store_name 顺序 → GT 1, 2)
actual_by_store = []
for r in asr:
    actual_by_store.append({
        "store_name": r.get("store_name"),
        "matched": [
            {
                "sku": s.get("matched_sku") or s.get("sku_code") or "",
                "unit_type": s.get("unit_type") or "",
                "product_name": s.get("product_name") or s.get("sku_name") or "",
                "match_score": s.get("match_score") or 0,
            }
            for s in (r.get("sku_results") or [])
        ],
    })

print("  GT vs Actual:")
for gt_key in ["1", "2"]:
    gt_skus = gt.get(gt_key, [])
    actual = actual_by_store[int(gt_key) - 1] if int(gt_key) - 1 < len(actual_by_store) else {"matched": []}
    actual_skus = actual["matched"]
    print(f"\n  ── GT Store #{gt_key} ({actual.get('store_name')}) ──")
    print(f"    GT count: {len(gt_skus)}, Actual matched: {len(actual_skus)}")
    
    # 按位置对比
    max_len = max(len(gt_skus), len(actual_skus))
    correct = 0
    for i in range(max_len):
        gt_item = gt_skus[i] if i < len(gt_skus) else None
        actual_item = actual_skus[i] if i < len(actual_skus) else None
        if gt_item and actual_item:
            gt_sku = gt_item.get("sku", "")
            act_sku = actual_item.get("sku", "")
            gt_ut = gt_item.get("unit_type", "")
            act_ut = actual_item.get("unit_type", "")
            match_sku = (gt_sku == act_sku)
            match_ut = (gt_ut == act_ut)
            mark = "✅" if match_sku else "❌"
            ut_mark = "✅" if match_ut else "❌"
            if match_sku:
                correct += 1
            print(f"    #{i+1}: {mark} GT={gt_sku} ({gt_ut}) | Actual={act_sku} ({act_ut}) score={actual_item.get('match_score', 0):.2f} name={actual_item.get('product_name','')[:30]!r}")
        elif gt_item and not actual_item:
            print(f"    #{i+1}: ⚠️ GT={gt_item.get('sku')} but no Actual match")
        elif actual_item and not gt_item:
            print(f"    #{i+1}: ⚠️ Extra Actual={actual_item.get('sku')} not in GT")
    
    total_gt = len(gt_skus)
    accuracy = correct / total_gt * 100 if total_gt else 0
    print(f"    📊 Store #{gt_key} 准确率: {correct}/{total_gt} = {accuracy:.1f}%")

total_gt_all = sum(len(gt.get(k, [])) for k in ["1", "2"])
total_correct = 0
for gt_key in ["1", "2"]:
    gt_skus = gt.get(gt_key, [])
    actual = actual_by_store[int(gt_key) - 1] if int(gt_key) - 1 < len(actual_by_store) else {"matched": []}
    actual_skus = actual["matched"]
    for i in range(min(len(gt_skus), len(actual_skus))):
        if gt_skus[i].get("sku") == actual_skus[i].get("sku"):
            total_correct += 1
overall_acc = total_correct / total_gt_all * 100 if total_gt_all else 0
print(f"\n  📊 总体准确率 (SKU match): {total_correct}/{total_gt_all} = {overall_acc:.1f}%")

print()
print("=" * 80)
print("✅ 报告完成")
print("=" * 80)
